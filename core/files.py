"""Offline file naming and filename spreadsheet business logic.

The module deliberately keeps filesystem operations separate from the Tk UI so
the rules can be exercised in tests and reused by a future packaged build.
"""

from __future__ import annotations

import ctypes
import hashlib
import json
import mimetypes
import os
import re
import subprocess
import uuid
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable, Sequence

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - UI reports this clearly
    Workbook = None
    load_workbook = None


ILLEGAL_CHARS = set('\\/:*?"<>|')
RESERVED_NAMES = {
    "CON", "PRN", "AUX", "NUL",
    *(f"COM{i}" for i in range(1, 10)),
    *(f"LPT{i}" for i in range(1, 10)),
}


def natural_key(value: str):
    """Sort text like a human: file2 comes before file10."""
    return [int(part) if part.isdigit() else part.casefold()
            for part in re.split(r"(\d+)", value)]


def normalise_ext(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    return value if value.startswith(".") else f".{value}"


def read_file_metadata(path: str | Path, root: str | Path | None = None) -> dict[str, Any]:
    """Return metadata that is meaningful for every file type.

    Format-specific namespaces are deliberately supplied by workflow providers
    through ``scan_folder(..., metadata_reader=...)``.
    """
    source = Path(path)
    stat = source.stat()
    mime_type, _encoding = mimetypes.guess_type(source.name)
    metadata: dict[str, Any] = {
        "file": {
            "name": source.name,
            "stem": source.stem,
            "extension": source.suffix.lstrip(".").casefold(),
            "size_bytes": stat.st_size,
            "created_ns": stat.st_ctime_ns,
            "created_date": datetime.fromtimestamp(stat.st_ctime).strftime("%Y%m%d"),
            "modified_ns": stat.st_mtime_ns,
            "mime_type": mime_type or "application/octet-stream",
        }
    }
    if root is not None:
        try:
            metadata["file"]["relative_path"] = str(source.relative_to(Path(root)))
        except ValueError:
            pass
    return metadata


def is_generated_workbook(path: Path) -> bool:
    """Generated exports, including .oriNN, are never scanned again."""
    return bool(re.fullmatch(r".+(?:\.ori\d+)?\.ffnf\.xlsx", path.name, re.I))


def is_generated_structure(path: Path) -> bool:
    """The structure companion is an export artifact, not source content."""
    return path.name.casefold() == "filetree.txt"


def _windows_file_attributes(path: Path) -> int:
    if os.name != "nt":
        return 0
    try:
        return ctypes.windll.kernel32.GetFileAttributesW(str(path))
    except Exception:
        return 0


def is_hidden(path: Path) -> bool:
    attrs = _windows_file_attributes(path)
    return bool(attrs != 0xFFFFFFFF and attrs & 0x2) or path.name.startswith(".")


def is_system(path: Path) -> bool:
    attrs = _windows_file_attributes(path)
    return bool(attrs != 0xFFFFFFFF and attrs & 0x4)


@dataclass
class FileRecord:
    path: str
    root: str
    extension: str
    folder_name: str
    relative_folder: str
    original_name: str
    stem: str
    selected: bool = True
    child_prefix: str = ""
    name: str = ""
    base_name: str = ""
    status: str = "Ready"
    status_detail: str = ""
    target_name: str = ""
    # The unsuffixed workflow target is retained so conflict disambiguation
    # can be recomputed after a preview refresh without stacking `_01` again.
    workflow_base_target_name: str = ""
    excel_source: str = ""
    group_key: str = ""
    extension_original: str = ""
    removed: bool = False
    parsed_fields: dict[str, str] = field(default_factory=dict)
    metadata: dict[str, Any] = field(default_factory=dict)
    parse_unmatched: str = ""
    parse_confidence: float = 0.0
    parse_error: str = ""
    association_id: str = ""
    associated_extensions: list[str] = field(default_factory=list)
    # Values entered through the active declarative workflow.  They are kept
    # separate from the legacy fields so the default workflow remains fully
    # compatible with existing imports and undo history.
    workflow_values: dict[str, str] = field(default_factory=dict)
    workflow_candidates: dict[str, list[str]] = field(default_factory=dict)
    workflow_candidate_details: dict[str, list[dict[str, str]]] = field(default_factory=dict)
    workflow_derived: dict[str, Any] = field(default_factory=dict)
    workflow_actions: set[str] = field(default_factory=set)
    workflow_auto_fields: set[str] = field(default_factory=set)
    workflow_number_fields: set[str] = field(default_factory=set)
    workflow_manual_fields: set[str] = field(default_factory=set)

    def __post_init__(self):
        if not self.name:
            self.name = self.stem
        if not self.base_name:
            self.base_name = self.name
        if not self.extension_original:
            self.extension_original = self.extension
        if not self.group_key:
            self.group_key = f"{self.relative_folder}\x1f{self.extension.casefold()}"

    @property
    def source_path(self) -> Path:
        return Path(self.path)

    @property
    def group_label(self) -> str:
        return f"{self.folder_name} / {self.extension.lstrip('.').upper() or '无扩展名'}"


@dataclass
class NamingGroup:
    key: str
    folder: str
    folder_name: str
    extension: str
    records: list[FileRecord] = field(default_factory=list)
    selected: bool = True
    prefix: str = ""
    relative_folder: str = ""
    meta_prefix: str = ""
    workflow_values: dict[str, str] = field(default_factory=dict)
    workflow_candidates: dict[str, list[str]] = field(default_factory=dict)
    # A workflow may combine several extensions into one logical group. The
    # legacy ``extension`` value remains for compatibility with single-format
    # groups, while this list drives the mixed-format label and API payload.
    extensions: list[str] = field(default_factory=list)

    @property
    def label(self) -> str:
        folder = self.relative_folder or self.folder_name
        extension_label = "图像" if self.extension == ".image" else self.extension.lstrip(".").upper() or "无扩展名"
        return f"{folder} / {extension_label} ({len(self.records)})"


def directory_prefix_defaults(root: str | Path, directory: str | Path,
                              mapping: dict[str, int | None] | None = None) -> tuple[str, str, str]:
    """Return the scheme-one defaults for a file-containing directory.

    The three fields are the last three directory names, with the root name
    included when a shallow path does not yet contain three descendants.
    For example, ``Root/A/B/C`` becomes ``A, B, C`` while ``Root/A/B``
    becomes ``Root, A, B``.  A file directly in the root has an empty group
    and child prefix so the root name is not duplicated by default.
    """
    root_path = Path(root).expanduser().resolve()
    directory_path = Path(directory).expanduser().resolve()
    try:
        relative_parts = directory_path.relative_to(root_path).parts
    except ValueError:
        relative_parts = directory_path.parts
    names = [root_path.name or str(root_path)] + list(relative_parts)
    if mapping is not None:
        def mapped(field: str) -> str:
            value = mapping.get(field)
            if value is None:
                return ""
            try:
                index = int(value)
            except (TypeError, ValueError):
                return ""
            if index < 0:
                index = len(names) + index
            if 0 <= index < len(names):
                return names[index]
            return ""
        return mapped("meta"), mapped("group"), mapped("child")
    if len(names) >= 3:
        return tuple(names[-3:])  # type: ignore[return-value]
    if len(names) == 2:
        return names[0], names[1], ""
    return names[0] if names else "", "", ""


@dataclass
class ScanResult:
    root: str
    records: list[FileRecord]
    extension_counts: dict[str, int]
    groups: dict[str, NamingGroup]
    skipped: list[str] = field(default_factory=list)
    associations: list[dict[str, Any]] = field(default_factory=list)
    max_depth: int = 0


@dataclass
class LogEntry:
    level: str
    message: str
    timestamp: str = field(default_factory=lambda: datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds"))


@dataclass
class RenameItem:
    old_path: str
    new_path: str
    group_key: str
    success: bool
    error: str = ""
    old_fingerprint: dict[str, Any] = field(default_factory=dict)
    new_fingerprint: dict[str, Any] = field(default_factory=dict)


@dataclass
class RenameOperation:
    operation_time: str
    kind: str
    items: list[RenameItem]
    transaction_status: str = "committed"
    transaction_error: str = ""


@dataclass
class ValidationIssue:
    record: FileRecord
    code: str
    message: str


@dataclass
class ExcelMatchResult:
    mode: str
    mapping: dict[str, str]
    matched_count: int
    unmatched_files: list[FileRecord]
    unmatched_rows: list[tuple[int, str, str]]
    warnings: list[str]
    matched_without_name: list[FileRecord] = field(default_factory=list)
    sheet_name: str = ""
    detail_mode: bool = False


def scan_folder(root: str | Path, include_hidden: bool = False,
                include_system: bool = False,
                directory_mapping: dict[str, int | None] | None = None,
                metadata_reader: Any | None = None) -> ScanResult:
    root_path = Path(root).expanduser().resolve()
    if not root_path.is_dir():
        raise NotADirectoryError(str(root_path))
    records: list[FileRecord] = []
    skipped: list[str] = []
    counts: dict[str, int] = {}
    groups: dict[str, NamingGroup] = {}
    max_depth = 0
    for directory, dirs, files in os.walk(root_path):
        dirs[:] = sorted(dirs, key=natural_key)
        directory_path = Path(directory)
        if not include_hidden:
            dirs[:] = [d for d in dirs if not is_hidden(directory_path / d)]
        if not include_system:
            dirs[:] = [d for d in dirs if not is_system(directory_path / d)]
        for filename in sorted(files, key=natural_key):
            path = directory_path / filename
            if is_generated_workbook(path) or is_generated_structure(path):
                skipped.append(str(path))
                continue
            if not include_hidden and is_hidden(path):
                continue
            if not include_system and is_system(path):
                continue
            try:
                if not path.is_file():
                    continue
            except OSError as exc:
                skipped.append(f"{path}: {exc}")
                continue
            original_ext = path.suffix
            ext = original_ext.lower()
            relative_folder = os.path.relpath(directory_path, root_path)
            if relative_folder == ".":
                relative_folder = "(root)"
            folder_name = directory_path.name or str(directory_path)
            try:
                max_depth = max(max_depth, len(directory_path.relative_to(root_path).parts))
            except ValueError:
                pass
            default_meta, default_group, default_child = directory_prefix_defaults(root_path, directory_path, directory_mapping)
            rec = FileRecord(
                path=str(path), root=str(root_path), extension=ext,
                folder_name=folder_name, relative_folder=relative_folder,
                original_name=filename, stem=path.stem,
                extension_original=original_ext,
                child_prefix=default_child,
            )
            try:
                reader = metadata_reader or read_file_metadata
                rec.metadata = reader(path, root_path)
            except (OSError, ValueError, TypeError) as exc:
                rec.metadata = {"file": {"error": str(exc)}}
            records.append(rec)
            counts[ext] = counts.get(ext, 0) + 1
            group = groups.get(rec.group_key)
            if group is None:
                group = NamingGroup(
                    rec.group_key,
                    str(directory_path),
                    folder_name,
                    ext,
                    prefix=default_group,
                    relative_folder=relative_folder,
                    meta_prefix=default_meta,
                    extensions=[ext],
                )
                groups[rec.group_key] = group
            group.records.append(rec)
    for group in groups.values():
        group.records.sort(key=lambda item: natural_key(item.original_name))
    records.sort(key=lambda item: (natural_key(item.relative_folder), natural_key(item.original_name)))
    associations = refresh_stem_associations(records)
    return ScanResult(str(root_path), records, dict(sorted(counts.items(), key=lambda pair: natural_key(pair[0]))), groups, skipped, associations, max_depth)


def build_stem_associations(records: Sequence[FileRecord]) -> list[dict[str, Any]]:
    """Group files in one directory that share a stem across extensions.

    Each extension remains in its own naming group.  The WebUI may use the
    resulting identity to synchronize a shared target stem.
    """
    buckets: dict[tuple[str, str], list[FileRecord]] = {}
    for record in records:
        if record.removed:
            continue
        key = (record.relative_folder.casefold(), record.stem.casefold())
        buckets.setdefault(key, []).append(record)
    result: list[dict[str, Any]] = []
    for (relative_folder, stem), items in buckets.items():
        extensions = sorted({item.extension for item in items}, key=natural_key)
        if len(extensions) < 2:
            continue
        association_id = f"{relative_folder}\x1f{stem}"
        result.append({
            "id": association_id,
            "relative_folder": relative_folder,
            "stem": items[0].stem,
            "extensions": extensions,
            "paths": [item.path for item in sorted(items, key=lambda item: natural_key(item.extension))],
            "count": len(items),
        })
    return sorted(result, key=lambda item: (natural_key(item["relative_folder"]), natural_key(item["stem"])))


def refresh_stem_associations(records: Sequence[FileRecord]) -> list[dict[str, Any]]:
    """Rebuild association fields after records have been renamed in memory."""
    for record in records:
        record.association_id = ""
        record.associated_extensions = []
    associations = build_stem_associations(records)
    by_path = {record.path: record for record in records}
    for association in associations:
        for path in association["paths"]:
            record = by_path.get(path)
            if record:
                record.association_id = association["id"]
                record.associated_extensions = list(association["extensions"])
    return associations


def compose_filename(meta_prefix: str, group_prefix: str, child_prefix: str,
                     name: str, extension: str, separator: str = "_") -> str:
    # Strip only for the emptiness decision; legal leading/trailing spaces are
    # preserved so a user's root prefix is not silently changed in preview.
    fields = [str(value) for value in (meta_prefix, group_prefix, child_prefix, name)
              if str(value).strip()]
    stem = separator.join(fields) if fields else "unnamed"
    return f"{stem}{normalise_ext(extension)}"


_TEMPLATE_FIELDS = {
    "name": r"(?P<name>.+?)",
    "stem": r"(?P<name>.+?)",
    "number": r"(?P<number>\d{1,5})",
    "type": r"(?P<type>.+?)",
    "category": r"(?P<category>.+?)",
    "pack": r"(?P<pack>.+?)",
    "*": r"(?P<unmatched>.+?)",
}


def _template_regex(template: str, field_patterns: dict[str, str] | None = None) -> tuple[re.Pattern[str] | None, list[str]]:
    """Compile a simple ``{field}`` filename template into a full regex."""
    patterns = dict(_TEMPLATE_FIELDS)
    if field_patterns:
        patterns.update({str(key).casefold(): str(value) for key, value in field_patterns.items()})
    if not template or template.strip().casefold() in {"auto", "自动"}:
        return None, []
    parts: list[str] = []
    fields: list[str] = []
    cursor = 0
    for match in re.finditer(r"\{([\w*]+)\}", template):
        parts.append(re.escape(template[cursor:match.start()]))
        field_name = match.group(1).casefold()
        pattern = patterns.get(field_name)
        if pattern is None:
            # Unknown placeholders are treated as literal text rather than
            # silently accepting a malformed template.
            return None, []
        group_name = field_name
        if field_name == "stem":
            group_name = "name"
        if group_name in fields:
            return None, []
        fields.append(group_name)
        parts.append(pattern.replace(f"?P<{group_name}>", f"?P<{group_name}>"))
        cursor = match.end()
    parts.append(re.escape(template[cursor:]))
    try:
        return re.compile("^" + "".join(parts) + "$", re.IGNORECASE), fields
    except re.error:
        return None, []


def parse_filename(stem: str, template: str = "auto",
                   field_patterns: dict[str, str] | None = None) -> dict[str, Any]:
    """Parse a filename stem for preview and optional name generation.

    The core only knows generic naming tokens.  A workflow may provide extra
    field patterns through a module-owned parser before calling this function.
    ``auto`` recognizes a trailing number without treating every number as a
    sequence number.
    """
    source = str(stem or "")
    auto_mode = not template or template.strip().casefold() in {"auto", "自动"}
    regex, fields = _template_regex(template, field_patterns)
    if not auto_mode and regex is None:
        return {"stem": source, "fields": {}, "unmatched": source,
                "confidence": 0.0, "matched": False, "template": template,
                "error": "模板无效或包含重复/未知字段"}
    if regex is not None:
        matched = regex.fullmatch(source)
        if not matched:
            return {"stem": source, "fields": {}, "unmatched": source,
                    "confidence": 0.0, "matched": False}
        values = {key: (value or "") for key, value in matched.groupdict().items()}
        unmatched = values.pop("unmatched", "")
        if "category" in values and "type" not in values:
            values["type"] = values["category"]
        return {"stem": source, "fields": values, "unmatched": unmatched,
                "confidence": 1.0, "matched": True, "template": template, "field_order": fields}

    working = source
    fields_out: dict[str, str] = {}
    number_match = re.search(r"(?:^|[_\-\s])([0-9]{1,5})$", working)
    if number_match:
        fields_out["number"] = number_match.group(1)
        working = working[:number_match.start()].rstrip(" _-.")
    tokens = [token for token in re.split(r"[_\-]+", working) if token]
    if tokens:
        if len(tokens) > 1:
            fields_out["type"] = tokens[0]
            fields_out["name"] = "_".join(tokens[1:])
        else:
            fields_out["name"] = tokens[0]
    elif source:
        fields_out["name"] = source
    recognized = sum(bool(value) for value in fields_out.values())
    confidence = min(1.0, recognized / 4.0) if source else 0.0
    return {"stem": source, "fields": fields_out, "unmatched": "",
            "confidence": confidence, "matched": bool(fields_out), "template": "auto",
            "field_order": list(fields_out)}


def apply_filename_parse(records: Sequence[FileRecord], template: str = "auto",
                         use_name: bool = False,
                         parser: Callable[[str, str], dict[str, Any]] | None = None) -> None:
    for record in records:
        previous_parsed_name = record.parsed_fields.get("name", "")
        previous_name = record.name
        parsed = parser(record.stem, template) if parser else parse_filename(record.stem, template)
        record.parsed_fields = dict(parsed.get("fields", {}))
        record.parse_unmatched = str(parsed.get("unmatched", ""))
        record.parse_confidence = float(parsed.get("confidence", 0.0) or 0.0)
        record.parse_error = str(parsed.get("error", ""))
        if use_name:
            record.name = record.parsed_fields.get("name") or record.stem
        elif previous_parsed_name and previous_name == previous_parsed_name:
            record.name = record.stem


def preview_group(group: NamingGroup, meta_prefix: str, separator: str = "_") -> None:
    for record in group.records:
        record.target_name = compose_filename(meta_prefix, group.prefix, record.child_prefix,
                                               record.name, record.extension_original or record.extension, separator)
        record.status = "Ready" if record.selected else "Skipped"
        record.status_detail = "" if record.selected else "Not selected"


def assign_numeric(group: NamingGroup, start: int = 1, width: int = 2,
                   step: int = 1, meta_prefix: str = "", separator: str = "_") -> None:
    current = start
    for record in group.records:
        if record.selected and record.status not in {"Conflict", "Error", "Skipped", "未匹配"}:
            record.name = str(current).zfill(max(1, width))
            current += step
            record.target_name = compose_filename(meta_prefix, group.prefix, record.child_prefix,
                                               record.name, record.extension_original or record.extension, separator)


def _path_case_key(path: Path) -> str:
    return os.path.normcase(os.path.abspath(str(path)))


def validate_filename(name: str) -> str | None:
    if not name or name in {".", ".."}:
        return "文件名不能为空"
    illegal = sorted({char for char in name if char in ILLEGAL_CHARS})
    if illegal:
        return f"包含 Windows 非法字符: {' '.join(illegal)}"
    if any(ord(char) < 32 for char in name):
        return "包含控制字符"
    if name.endswith((" ", ".")):
        return "文件名不能以空格或点结尾"
    # Windows reserves the device token before the first dot (CON.txt is
    # still invalid), regardless of the rest of the stem.
    base = name.split(".", 1)[0].upper()
    if base in RESERVED_NAMES:
        return f"保留设备名不可用: {base}"
    return None


def preflight(records: Sequence[FileRecord], separator: str = "_") -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    targets: dict[str, FileRecord] = {}
    selected_sources = {
        _path_case_key(record.source_path)
        for record in records
        if record.selected
    }
    for record in records:
        if not record.selected:
            continue
        if record.status == "Conflict" and record.status_detail.startswith("缺少工作流必填字段"):
            issues.append(ValidationIssue(record, "workflow_required", record.status_detail))
        source = record.source_path
        target_name = record.target_name or compose_filename("", "", record.child_prefix, record.name,
                                                            record.extension_original or record.extension, separator)
        target = source.with_name(target_name)
        record.target_name = target_name
        if not source.exists():
            issues.append(ValidationIssue(record, "missing", "源文件不存在"))
            continue
        error = validate_filename(target_name)
        if error:
            issues.append(ValidationIssue(record, "invalid", error))
        if len(str(target)) > 260:
            issues.append(ValidationIssue(record, "too_long", "路径或文件名超过 Windows 传统长度限制"))
        target_key = _path_case_key(target)
        source_key = _path_case_key(source)
        previous = targets.get(target_key)
        if previous is not None and _path_case_key(previous.source_path) != source_key:
            issues.append(ValidationIssue(record, "duplicate", f"目标名称与 {previous.original_name} 重复"))
        else:
            targets[target_key] = record
        if target_key != source_key and target.exists() and target_key not in selected_sources:
            issues.append(ValidationIssue(record, "exists", "目标文件已存在，不覆盖已有文件"))
        try:
            with source.open("rb"):
                pass
        except PermissionError:
            issues.append(ValidationIssue(record, "permission", "没有访问权限或文件被占用"))
        except OSError as exc:
            issues.append(ValidationIssue(record, "io", f"无法访问源文件: {exc}"))
    return issues


def file_fingerprint(path: str | Path, include_hash: bool = True) -> dict[str, Any]:
    """Return a durable identity snapshot used to guard undo/redo.

    The SHA-256 is intentionally included by default so a replacement file at
    the same path cannot be mistaken for the renamed source.  Stat metadata is
    retained as a quick diagnostic and for older/large-file callers that opt
    out of hashing.
    """
    source = Path(path)
    stat = source.stat()
    result: dict[str, Any] = {
        "size": int(stat.st_size),
        "mtime_ns": int(getattr(stat, "st_mtime_ns", int(stat.st_mtime * 1_000_000_000))),
        "file_id": int(getattr(stat, "st_ino", 0) or 0),
    }
    if include_hash:
        digest = hashlib.sha256()
        with source.open("rb") as stream:
            for chunk in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(chunk)
        result["sha256"] = digest.hexdigest()
    return result


def _fingerprint_matches(path: Path, expected: dict[str, Any]) -> bool:
    if not expected:
        return True  # legacy history written before fingerprints existed
    try:
        actual = file_fingerprint(path, include_hash=bool(expected.get("sha256")))
    except OSError:
        return False
    if expected.get("sha256"):
        if (actual.get("sha256") != expected.get("sha256")
                or (expected.get("size") is not None and actual.get("size") != expected.get("size"))):
            return False
        expected_file_id = int(expected.get("file_id", 0) or 0)
        actual_file_id = int(actual.get("file_id", 0) or 0)
        return not expected_file_id or not actual_file_id or expected_file_id == actual_file_id
    return all(actual.get(key) == value for key, value in expected.items() if key in actual)


def _resolve_history_source(expected_path: Path, destination_path: Path,
                            expected_fingerprint: dict[str, Any]) -> tuple[Path | None, str]:
    """Resolve an undo/redo source by identity, using paths only as hints."""
    hints: list[Path] = []
    for hint in (expected_path, destination_path):
        if _path_case_key(hint) not in {_path_case_key(item) for item in hints}:
            hints.append(hint)
    for hint in hints:
        if hint.is_file() and _fingerprint_matches(hint, expected_fingerprint):
            return hint, ""
    if not expected_fingerprint:
        return None, f"历史记录没有文件指纹，且预期路径不存在: {expected_path}"

    matches: list[tuple[Path, dict[str, Any]]] = []
    expected_size = expected_fingerprint.get("size")
    searched: set[str] = set()
    for directory in (expected_path.parent, destination_path.parent):
        directory_key = _path_case_key(directory)
        if directory_key in searched or not directory.is_dir():
            continue
        searched.add(directory_key)
        try:
            children = list(directory.iterdir())
        except OSError:
            continue
        for candidate in children:
            try:
                if not candidate.is_file() or ".ffnf-txn-" in candidate.name:
                    continue
                if expected_size is not None and candidate.stat().st_size != int(expected_size):
                    continue
                actual = file_fingerprint(candidate, include_hash=bool(expected_fingerprint.get("sha256")))
            except (OSError, TypeError, ValueError):
                continue
            if expected_fingerprint.get("sha256"):
                matched = actual.get("sha256") == expected_fingerprint.get("sha256")
            else:
                matched = all(actual.get(key) == value for key, value in expected_fingerprint.items()
                              if key in actual)
            if matched:
                matches.append((candidate, actual))

    expected_file_id = int(expected_fingerprint.get("file_id", 0) or 0)
    if expected_file_id:
        same_id = [candidate for candidate, actual in matches
                   if int(actual.get("file_id", 0) or 0) == expected_file_id]
        if len(same_id) == 1:
            return same_id[0], ""
        if len(same_id) > 1:
            names = "、".join(str(path) for path in same_id[:3])
            return None, f"文件指纹匹配到多个候选，无法安全识别: {names}"
    if len(matches) == 1:
        return matches[0][0], ""
    if len(matches) > 1:
        names = "、".join(str(path) for path, _actual in matches[:3])
        return None, f"文件内容指纹匹配到多个候选，无法安全识别: {names}"
    return None, f"未在原目录找到匹配文件指纹的文件: {expected_path.parent}"


def _move_history_items(items: Sequence[dict[str, Any]], direction: str) -> list[str]:
    """Move one history action transactionally after resolving every file."""
    plans: list[tuple[dict[str, Any], Path, Path]] = []
    already_at_destination: list[tuple[dict[str, Any], Path, Path]] = []
    errors: list[str] = []
    for item in items:
        try:
            old_path = Path(item["old_path"])
            new_path = Path(item["new_path"])
        except (KeyError, TypeError, ValueError) as exc:
            errors.append(f"历史路径无效: {exc}")
            continue
        if direction == "undo":
            expected_path, destination = new_path, old_path
            fingerprint = item.get("new_fingerprint", {})
        else:
            expected_path, destination = old_path, new_path
            fingerprint = item.get("undo_fingerprint") or item.get("old_fingerprint", {})
        source, error = _resolve_history_source(expected_path, destination, fingerprint)
        if source is None:
            errors.append(error)
            continue
        if str(source) == str(destination):
            already_at_destination.append((item, source, destination))
        else:
            plans.append((item, source, destination))
    if errors:
        return errors

    source_keys: dict[str, Path] = {}
    destination_keys: dict[str, Path] = {}
    for _item, source, destination in plans:
        source_key = _path_case_key(source)
        destination_key = _path_case_key(destination)
        if source_key in source_keys:
            errors.append(f"多个历史项目解析到同一文件指纹: {source}")
        else:
            source_keys[source_key] = source
        if destination_key in destination_keys:
            errors.append(f"多个历史项目要求同一目标路径: {destination}")
        else:
            destination_keys[destination_key] = destination
    for _item, _source, destination in plans:
        destination_key = _path_case_key(destination)
        if destination.exists() and destination_key not in source_keys:
            errors.append(f"目标路径已存在，不覆盖: {destination}")
    if errors:
        return errors

    staged: list[tuple[dict[str, Any], Path, Path, Path]] = []
    committed: list[tuple[dict[str, Any], Path, Path]] = []
    try:
        for item, source, destination in plans:
            temporary = source.with_name(f".{source.name}.ffnf-txn-{uuid.uuid4().hex}.tmp")
            os.rename(source, temporary)
            staged.append((item, source, destination, temporary))
        for item, source, destination, temporary in staged:
            os.rename(temporary, destination)
            committed.append((item, source, destination))
    except Exception as exc:
        rollback_errors: list[str] = []
        for _item, source, destination in reversed(committed):
            try:
                same_path_different_case = (_path_case_key(destination) == _path_case_key(source)
                                            and str(destination) != str(source))
                if destination.exists() and (same_path_different_case or not source.exists()):
                    if same_path_different_case:
                        rollback_temporary = destination.with_name(
                            f".{destination.name}.ffnf-txn-{uuid.uuid4().hex}.tmp"
                        )
                        os.rename(destination, rollback_temporary)
                        os.rename(rollback_temporary, source)
                    else:
                        os.rename(destination, source)
            except Exception as rollback_exc:
                rollback_errors.append(f"{destination} -> {source}: {rollback_exc}")
        committed_ids = {id(item) for item, _source, _destination in committed}
        for item, source, _destination, temporary in reversed(staged):
            if id(item) in committed_ids:
                continue
            try:
                if temporary.exists() and not source.exists():
                    os.rename(temporary, source)
            except Exception as rollback_exc:
                rollback_errors.append(f"{temporary} -> {source}: {rollback_exc}")
        detail = f"事务执行失败，已回滚: {exc}"
        if rollback_errors:
            detail += "；部分回滚失败: " + "；".join(rollback_errors)
        return [detail]

    for item, source, destination in already_at_destination + committed:
        try:
            fingerprint = file_fingerprint(destination)
        except OSError:
            fingerprint = {}
        if direction == "undo":
            item["undone"] = True
            item["undo_source_path"] = str(source)
            item["undo_fingerprint"] = fingerprint
        else:
            item["undone"] = False
            item["redo_source_path"] = str(source)
            item["new_fingerprint"] = fingerprint
            item.pop("redo_error", None)
    return []


def append_history(history_path: str | Path, operation: RenameOperation) -> None:
    path = Path(history_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    data = []
    if path.exists():
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            data = []
    data.append({
        "operation_time": operation.operation_time,
        "kind": operation.kind,
        "items": [asdict(item) for item in operation.items],
    })
    # Replace the history atomically so an interrupted write cannot leave a
    # truncated JSON file and make the durable undo stack unusable.
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    temporary.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temporary, path)


def execute_rename(records: Sequence[FileRecord], history_path: str | Path,
                   kind: str = "batch", separator: str = "_",
                   write_history: bool = True) -> RenameOperation:
    selected = [record for record in records if record.selected]
    issues = preflight(selected, separator)
    issue_by_path: dict[str, list[ValidationIssue]] = {}
    for issue in issues:
        issue_by_path.setdefault(issue.record.path, []).append(issue)
        issue.record.status = "Conflict"
        issue.record.status_detail = issue.message
    if issues:
        operation = RenameOperation(datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds"), kind, [], "blocked", "预检查发现冲突，事务未执行")
        for record in selected:
            if record.path in issue_by_path:
                operation.items.append(RenameItem(record.path, str(record.source_path.with_name(record.target_name)), record.group_key, False,
                                                  "; ".join(issue.message for issue in issue_by_path[record.path])))
        # A blocked preflight has no filesystem mutation, so it is not an
        # undoable operation and must not hide the previous successful entry.
        return operation
    operation_time = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    items: list[RenameItem] = []
    moving: list[tuple[FileRecord, Path, Path, dict[str, Any]]] = []
    for record in selected:
        source = record.source_path
        target = source.with_name(record.target_name)
        try:
            old_fingerprint = file_fingerprint(source)
        except OSError as exc:
            record.status = "Error"
            record.status_detail = str(exc)
            items.append(RenameItem(str(source), str(target), record.group_key, False, str(exc)))
            continue
        if str(source) == str(target):
            record.status = "Unchanged"
            items.append(RenameItem(str(source), str(target), record.group_key, True,
                                    old_fingerprint=old_fingerprint, new_fingerprint=old_fingerprint))
        else:
            moving.append((record, source, target, old_fingerprint))

    staged: list[tuple[FileRecord, Path, Path, Path, dict[str, Any]]] = []
    committed: list[tuple[FileRecord, Path, Path, dict[str, Any]]] = []
    transaction_error = ""
    rollback_failed = False
    stage_failed_record: tuple[FileRecord, Path, Path] | None = None
    # Stage every source under a unique sibling name first. This makes swaps
    # and chains safe and gives us a known point from which to compensate.
    try:
        for record, source, target, old_fingerprint in moving:
            temporary = source.with_name(f".{source.name}.ffnf-txn-{uuid.uuid4().hex}.tmp")
            stage_failed_record = (record, source, target)
            os.rename(source, temporary)
            staged.append((record, source, target, temporary, old_fingerprint))
        for record, source, target, temporary, old_fingerprint in staged:
            os.rename(temporary, target)
            new_fingerprint = file_fingerprint(target)
            committed.append((record, source, target, new_fingerprint))
            record.path = str(target)
            record.original_name = target.name
            record.stem = target.stem
            record.status = "Renamed"
            record.status_detail = ""
            items.append(RenameItem(str(source), str(target), record.group_key, True,
                                    old_fingerprint=old_fingerprint, new_fingerprint=new_fingerprint))
    except Exception as exc:
        transaction_error = str(exc)
        # Compensate final names and any still-staged names. Never overwrite.
        for record, source, target, new_fingerprint in reversed(committed):
            try:
                same_path_different_case = (_path_case_key(target) == _path_case_key(source)
                                            and str(target) != str(source))
                if target.exists() and (same_path_different_case or not source.exists()):
                    if same_path_different_case:
                        rollback_temporary = target.with_name(
                            f".{target.name}.ffnf-txn-{uuid.uuid4().hex}.tmp"
                        )
                        os.rename(target, rollback_temporary)
                        os.rename(rollback_temporary, source)
                    else:
                        os.rename(target, source)
                record.path = str(source)
                record.original_name = source.name
                record.stem = source.stem
            except Exception as rollback_exc:
                rollback_failed = True
                transaction_error += f"；回滚失败 {target} -> {source}: {rollback_exc}"
        for record, source, target, temporary, old_fingerprint in reversed(staged):
            if any(item[0] is record for item in committed):
                continue
            try:
                if temporary.exists() and not source.exists():
                    os.rename(temporary, source)
            except Exception as rollback_exc:
                rollback_failed = True
                transaction_error += f"；回滚失败 {temporary} -> {source}: {rollback_exc}"
        committed_paths = {str(source): str(target) for _record, source, target, _fp in committed}
        for item in items:
            if item.success and item.old_path in committed_paths:
                # A successfully compensated item is no longer an executed
                # rename and must not become an undo candidate.
                source = Path(item.old_path)
                target = Path(item.new_path)
                if source.exists() and not target.exists():
                    item.success = False
                    item.error = f"事务失败，已回滚：{transaction_error}"
        for record, source, target, _new_fingerprint in committed:
            record.status = "Error"
            record.status_detail = f"事务失败，已回滚：{transaction_error}"
        for record, source, target, temporary, old_fingerprint in staged:
            if not any(item[0] is record for item in committed):
                record.status = "Error"
                record.status_detail = f"事务失败，已回滚：{transaction_error}"
        if stage_failed_record is not None and not any(item[0] is stage_failed_record[0] for item in staged):
            record, source, target = stage_failed_record
            record.status = "Error"
            record.status_detail = f"事务失败：{transaction_error}"
            items.append(RenameItem(str(source), str(target), record.group_key, False, transaction_error))
        # Add failed items for moving records that never reached the final path.
        committed_records = {id(item[0]) for item in committed}
        for record, source, target, _temporary, _old_fingerprint in staged:
            if id(record) not in committed_records:
                items.append(RenameItem(str(source), str(target), record.group_key, False, transaction_error))
        operation = RenameOperation(operation_time, kind, items,
                                    "partial" if rollback_failed else "rolled_back",
                                    transaction_error)
    else:
        operation = RenameOperation(operation_time, kind, items, "committed", "")
    if write_history and items and any(item.success for item in items):
        append_history(history_path, operation)
    return operation


def _history_action_indices(data: list[dict[str, Any]], target_index: int,
                            eligible) -> list[int]:
    target = data[target_index]
    if target.get("kind") != "batch":
        return [target_index]
    operation_time = target.get("operation_time")
    start = target_index
    while start > 0:
        candidate = data[start - 1]
        if candidate.get("kind") != "batch" or candidate.get("operation_time") != operation_time:
            break
        start -= 1
    end = target_index
    while end + 1 < len(data):
        candidate = data[end + 1]
        if candidate.get("kind") != "batch" or candidate.get("operation_time") != operation_time:
            break
        end += 1
    return [index for index in range(start, end + 1) if eligible(data[index])]


def _write_history(path: Path, data: list[dict[str, Any]]) -> None:
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    temporary.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temporary, path)


def undo_last(history_path: str | Path) -> tuple[bool, list[str]]:
    path = Path(history_path)
    if not path.exists():
        return False, ["没有可撤销的历史记录"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:
        return False, [f"历史记录读取失败: {exc}"]
    if not isinstance(data, list) or not data:
        return False, ["没有可撤销的历史记录"]

    def eligible(operation: dict[str, Any]) -> bool:
        return (not operation.get("undone_at")
                and any(item.get("success") and not item.get("undone")
                        for item in operation.get("items", [])))

    target_index = next((index for index in range(len(data) - 1, -1, -1)
                         if eligible(data[index])), None)
    if target_index is None:
        return False, ["没有可撤销的历史记录"]
    operation_indices = _history_action_indices(data, target_index, eligible)
    items = [item for index in operation_indices for item in data[index].get("items", [])
             if item.get("success") and not item.get("undone")]
    errors = _move_history_items(items, "undo")
    now = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    if errors:
        for index in operation_indices:
            data[index]["undo_attempted_at"] = now
        _write_history(path, data)
        return False, [f"撤销失败：{error}" for error in errors]

    next_sequence = max((int(operation.get("undo_sequence", 0) or 0)
                         for operation in data), default=0) + 1
    for index in operation_indices:
        operation = data[index]
        successful = [item for item in operation.get("items", []) if item.get("success")]
        operation["restored_count"] = sum(bool(item.get("undone")) for item in successful)
        if successful and all(item.get("undone") for item in successful):
            operation["undone_at"] = now
            operation["undo_sequence"] = next_sequence
            operation.pop("undo_attempted_at", None)
    _write_history(path, data)
    return True, []


def redo_last(history_path: str | Path) -> tuple[bool, list[str]]:
    path = Path(history_path)
    if not path.exists():
        return False, ["没有可还原的撤销记录"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:
        return False, [f"历史记录读取失败: {exc}"]
    if not isinstance(data, list) or not data:
        return False, ["没有可还原的撤销记录"]

    def eligible(operation: dict[str, Any]) -> bool:
        return (bool(operation.get("undone_at"))
                and any(item.get("success") and item.get("undone")
                        for item in operation.get("items", [])))

    candidates = [index for index in range(len(data)) if eligible(data[index])]
    if not candidates:
        return False, ["没有可还原的撤销记录"]
    target_index = max(candidates, key=lambda index: (
        int(data[index].get("undo_sequence", 0) or 0), index,
    ))
    operation_indices = _history_action_indices(data, target_index, eligible)
    items = [item for index in operation_indices for item in data[index].get("items", [])
             if item.get("success") and item.get("undone")]
    if not items:
        return False, ["没有可还原的撤销文件"]
    errors = _move_history_items(items, "redo")
    now = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    if errors:
        for index in operation_indices:
            data[index]["redo_attempted_at"] = now
        _write_history(path, data)
        return False, [f"还原失败：{error}" for error in errors]

    for index in operation_indices:
        operation = data[index]
        successful = [item for item in operation.get("items", []) if item.get("success")]
        operation["redone_count"] = sum(not item.get("undone") for item in successful)
        if successful and all(not item.get("undone") for item in successful):
            operation.pop("undone_at", None)
            operation.pop("undo_attempted_at", None)
            operation.pop("redo_attempted_at", None)
            operation["redone_at"] = now
    _write_history(path, data)
    return True, []


def _excel_value(value) -> str:
    return "" if value is None else str(value).strip()


def _match_key(value: str, extension: str = "") -> str:
    value = _excel_value(value)
    if not value:
        return ""
    # Accept a bare filename or a Windows/POSIX path pasted into Excel.
    value = re.split(r"[\\/]", value)[-1]
    suffix = normalise_ext(extension)
    if suffix and value.casefold().endswith(suffix.casefold()):
        value = value[: -len(suffix)]
    return value.casefold()


def _excel_name(value: str, extension: str) -> str:
    value = _excel_value(value)
    if not value:
        return ""
    suffix = normalise_ext(extension)
    if suffix and value.casefold().endswith(suffix.casefold()):
        return value[: -len(suffix)]
    return value


_EXCEL_NAME_TEMPLATE = re.compile(r"\{([a-zA-Z][a-zA-Z0-9_.-]*)\}")


def _expand_excel_name_template(value: str, record: FileRecord,
                                row_values: dict[str, str] | None = None,
                                value_expander: Callable[[str, FileRecord, dict[str, str]], str] | None = None) -> str:
    """Expand placeholders through a caller-supplied workflow policy.

    The core fallback only uses values from the spreadsheet header row. A
    workflow can provide ``value_expander`` to combine those values with its
    own metadata namespace and normalizers.
    """
    source = _excel_name(value, record.extension)
    values = {str(key).casefold(): _excel_value(item)
              for key, item in (row_values or {}).items()}
    if value_expander:
        return value_expander(source, record, values)
    if not _EXCEL_NAME_TEMPLATE.search(source):
        return source
    expanded = _EXCEL_NAME_TEMPLATE.sub(
        lambda match: values.get(match.group(1).casefold(), ""), source
    )
    expanded = re.sub(r"([ _.-])\1+", r"\1", expanded)
    return expanded.strip(" _-.")


def import_xlsx(xlsx_path: str | Path, group: NamingGroup,
                sheet_name: str | None = None,
                value_expander: Callable[[str, FileRecord, dict[str, str]], str] | None = None) -> ExcelMatchResult:
    if load_workbook is None:
        raise RuntimeError("需要安装 openpyxl 才能导入 XLSX")
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet = None
        if sheet_name:
            sheet = next((candidate for candidate in workbook.worksheets
                          if candidate.title.casefold() == str(sheet_name).casefold()), None)
        if sheet is None:
            wanted = group.extension.lstrip(".").casefold()
            sheet = next((candidate for candidate in workbook.worksheets
                          if candidate.title.casefold() == wanted), workbook.worksheets[0])
        selected_sheet_name = sheet.title
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        # read_only workbooks keep a zip handle open on Windows until explicitly closed.
        workbook.close()
    mapping: dict[str, str] = {}
    unmatched_rows: list[tuple[int, str, str]] = []
    warnings: list[str] = []
    matched_without_name: list[FileRecord] = []
    records_by_key = {_match_key(record.original_name, group.extension): record for record in group.records}
    unmatched_files = list(group.records)
    detail_mode = False
    headers: list[str] = []
    if rows:
        headers = [str(value or "").strip().casefold() for value in rows[0]]
        header_tokens = {"source", "sourcename", "source name", "原文件名", "源文件名", "newname", "new name", "新名称", "新文件名", "relativepath", "relative path", "相对路径"}
        if any(token in header_tokens for token in headers):
            detail_mode = True
            source_index = next((index for index, value in enumerate(headers)
                                 if value in {"source", "sourcename", "source name", "原文件名", "源文件名"}), 0)
            name_index = next((index for index, value in enumerate(headers)
                               if value in {"newname", "new name", "新名称", "新文件名", "name", "名称"}), 1 if len(headers) > 1 else None)
            data_rows = rows[1:]
        else:
            source_index, name_index, data_rows = 0, 1 if sheet.max_column >= 2 else None, rows
    else:
        source_index, name_index, data_rows = 0, None, []
    has_b = name_index is not None and (detail_mode or sheet.max_column >= 2)
    if has_b:
        mode = "source-and-name"
        for row_number, row in enumerate(data_rows, start=2 if detail_mode else 1):
            source = _excel_value(row[source_index] if len(row) > source_index else "")
            new_name = _excel_value(row[name_index] if name_index is not None and len(row) > name_index else "")
            if not source and not new_name:
                continue
            key = _match_key(source, group.extension)
            record = records_by_key.get(key)
            if not record:
                unmatched_rows.append((row_number, source, new_name))
                warnings.append(f"WARN Excel 第 {row_number} 行无法匹配文件: {source}")
                continue
            if not new_name:
                warnings.append(f"WARN Excel 第 {row_number} 行 B 列为空，跳过: {source}")
                if record in unmatched_files:
                    unmatched_files.remove(record)
                matched_without_name.append(record)
                continue
            row_values = {
                headers[index]: _excel_value(cell)
                for index, cell in enumerate(row)
                if index < len(headers) and headers[index]
            }
            new_name = _expand_excel_name_template(new_name, record, row_values, value_expander)
            if not new_name:
                warnings.append(f"WARN Excel 第 {row_number} 行 B 列模板展开后为空，跳过: {source}")
                if record in unmatched_files:
                    unmatched_files.remove(record)
                matched_without_name.append(record)
                continue
            if record.path in mapping:
                warnings.append(f"WARN Excel 第 {row_number} 行重复匹配: {source}")
                continue
            mapping[record.path] = new_name
            record.excel_source = source
            if record in unmatched_files:
                unmatched_files.remove(record)
    else:
        mode = "ordered-names"
        row_index = 0
        for row_number, row in enumerate(data_rows, start=2 if detail_mode else 1):
            value = _excel_value(row[0] if row else "")
            if not value:
                unmatched_rows.append((row_number, "", ""))
                warnings.append(f"WARN Excel 第 {row_number} 行为空")
                row_index += 1
                continue
            if row_index >= len(group.records):
                unmatched_rows.append((row_number, value, ""))
                warnings.append(f"WARN Excel 第 {row_number} 行超出当前命名组文件数")
                row_index += 1
                continue
            record = group.records[row_index]
            mapping[record.path] = _excel_name(value, group.extension)
            record.excel_source = value
            if record in unmatched_files:
                unmatched_files.remove(record)
            row_index += 1
        if len(group.records) > row_index:
            warnings.append(f"WARN {len(group.records) - row_index} 个文件未分配 Excel 名称")
    return ExcelMatchResult(mode, mapping, len(mapping), unmatched_files, unmatched_rows, warnings,
                            matched_without_name, selected_sheet_name, detail_mode)


def _unique_export_path(folder: Path) -> Path:
    base = folder.name or "root"
    candidate = folder.parent / f"{base}.ffnf.xlsx"
    if not candidate.exists():
        return candidate
    index = 1
    while True:
        candidate = folder.parent / f"{base}.ori{index:02d}.ffnf.xlsx"
        if not candidate.exists():
            return candidate
        index += 1


def _structure_directories(root: str | Path, include_hidden: bool = False,
                           include_system: bool = False) -> tuple[dict[Path, list[Path]], int]:
    """Collect a filtered directory tree and return its maximum relative depth."""
    root_path = Path(root).expanduser().resolve()
    children: dict[Path, list[Path]] = {}
    max_depth = 0
    for directory, dirs, _files in os.walk(root_path):
        directory_path = Path(directory)
        dirs[:] = sorted(dirs, key=natural_key)
        if not include_hidden:
            dirs[:] = [name for name in dirs if not is_hidden(directory_path / name)]
        if not include_system:
            dirs[:] = [name for name in dirs if not is_system(directory_path / name)]
        child_paths = [directory_path / name for name in dirs]
        children[directory_path] = child_paths
        for child in child_paths:
            try:
                depth = len(child.relative_to(root_path).parts)
            except ValueError:
                continue
            max_depth = max(max_depth, depth)
    children.setdefault(root_path, [])
    return children, max_depth


def _write_filetree_export(root: str | Path, generated_tables: dict[str, Path],
                           include_hidden: bool = False,
                           include_system: bool = False) -> Path | None:
    """Write an index of content directories that received an XLSX export."""
    root_path = Path(root).expanduser().resolve()
    _children, max_depth = _structure_directories(root_path, include_hidden, include_system)
    if max_depth < 3 or not generated_tables:
        return None
    output = root_path / "filetree.txt"
    if output.exists():
        return None

    lines = [
        f"根目录: {root_path}",
        "生成 .ffnf.xlsx 的内容目录:",
    ]
    ordered = sorted(
        generated_tables.items(),
        key=lambda pair: natural_key(os.path.relpath(pair[0], root_path)),
    )
    for folder_text, workbook in ordered:
        folder = Path(folder_text).resolve()
        try:
            relative_folder = folder.relative_to(root_path).as_posix() or "."
        except ValueError:
            relative_folder = os.path.relpath(folder, root_path).replace(os.sep, "/")
        relative_workbook = os.path.relpath(Path(workbook).resolve(), root_path).replace(os.sep, "/")
        lines.append(f"- {relative_folder}")
        lines.append(f"  XLSX: {relative_workbook}")
    output.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return output


def collect_directory_statistics(root: str | Path, records: Sequence[FileRecord] | None = None,
                                 include_hidden: bool = False, include_system: bool = False) -> dict[str, Any]:
    root_path = Path(root).expanduser().resolve()
    result = scan_folder(root_path, include_hidden, include_system) if records is None else None
    effective_records = list(records if records is not None else result.records)
    tree, _max_depth = _structure_directories(root_path, include_hidden, include_system)
    directories: set[str] = {str(path) for path in tree}
    per_folder: dict[str, dict[str, Any]] = {}
    extension_counts: dict[str, int] = {}
    for record in effective_records:
        folder = str(Path(record.path).parent)
        directories.add(folder)
        info = per_folder.setdefault(folder, {"path": folder, "relative_folder": record.relative_folder,
                                               "file_count": 0, "extensions": {}})
        info["file_count"] += 1
        ext = record.extension or "(none)"
        info["extensions"][ext] = info["extensions"].get(ext, 0) + 1
        extension_counts[ext] = extension_counts.get(ext, 0) + 1
    return {
        "root": str(root_path),
        "directory_count": len(directories),
        "file_count": len(effective_records),
        "content_directory_count": len(per_folder),
        "empty_directory_count": max(0, len(directories) - len(per_folder)),
        "extension_counts": dict(sorted(extension_counts.items(), key=lambda pair: natural_key(pair[0]))),
        "folders": sorted(per_folder.values(), key=lambda item: natural_key(item["path"])),
    }


def _flatten_metadata(value: Any, prefix: str = "") -> dict[str, Any]:
    flattened: dict[str, Any] = {}
    if isinstance(value, dict):
        for key, child in value.items():
            child_prefix = f"{prefix}.{key}" if prefix else str(key)
            flattened.update(_flatten_metadata(child, child_prefix))
    elif prefix:
        flattened[prefix] = value
    return flattened


def export_filename_tables(root: str | Path, selected_extensions: Iterable[str],
                           include_hidden: bool = False, include_system: bool = False,
                           metadata_reader: Callable[[str | Path, str | Path | None], dict[str, Any]] | None = None) -> list[Path]:
    if Workbook is None:
        raise RuntimeError("需要安装 openpyxl 才能导出 XLSX")
    selected = {normalise_ext(ext).casefold() for ext in selected_extensions}
    result = scan_folder(root, include_hidden=include_hidden, include_system=include_system,
                         metadata_reader=metadata_reader)
    by_folder: dict[str, list[FileRecord]] = {}
    for record in result.records:
        if record.extension.casefold() in selected:
            by_folder.setdefault(record.path and str(Path(record.path).parent), []).append(record)
    outputs: list[Path] = []
    generated_tables: dict[str, Path] = {}
    for folder_text, records in sorted(by_folder.items(), key=lambda pair: natural_key(pair[0])):
        folder = Path(folder_text)
        workbook = Workbook()
        workbook.remove(workbook.active)
        by_ext: dict[str, list[FileRecord]] = {}
        for record in records:
            by_ext.setdefault(record.extension.lower(), []).append(record)
        used_titles: set[str] = set()
        for ext, ext_records in sorted(by_ext.items(), key=lambda pair: natural_key(pair[0])):
            title = ext.lstrip(".").upper()[:31] or "NO_EXT"
            original_title = title
            suffix = 1
            while title in used_titles:
                title = f"{original_title[:28]}_{suffix}"
                suffix += 1
            used_titles.add(title)
            sheet = workbook.create_sheet(title)
            ordered_records = sorted(ext_records, key=lambda item: natural_key(item.stem))
            metadata_values = [_flatten_metadata(record.metadata) for record in ordered_records]
            metadata_keys = sorted({key for values in metadata_values for key in values}, key=natural_key)
            metadata_headers = [f"Metadata.{key}" for key in metadata_keys]
            sheet.append(["SourceName", "NewName", "RelativePath", "Folder", "Extension", "SizeBytes", "ModifiedTime", "Association", *metadata_headers])
            for record in ordered_records:
                try:
                    stat = Path(record.path).stat()
                    size, modified = int(stat.st_size), datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds")
                except OSError:
                    size, modified = "", ""
                values = _flatten_metadata(record.metadata)
                sheet.append([record.stem, "", record.relative_folder, record.folder_name,
                              record.extension, size, modified, record.association_id,
                              *[values.get(key, "") for key in metadata_keys]])
        stats = collect_directory_statistics(root, result.records, include_hidden, include_system)
        metadata = workbook.create_sheet("Metadata")
        metadata.append(["Field", "Value"])
        metadata.append(["Root", stats["root"]])
        metadata.append(["DirectoryCount", stats["directory_count"]])
        metadata.append(["FileCount", stats["file_count"]])
        metadata.append(["ContentDirectoryCount", stats["content_directory_count"]])
        metadata.append(["AssociationCount", len(result.associations)])
        summary = workbook.create_sheet("Summary")
        summary.append(["RelativeFolder", "Path", "FileCount", "Extensions"])
        for folder_info in stats["folders"]:
            extension_text = ", ".join(f"{key}:{value}" for key, value in folder_info["extensions"].items())
            summary.append([folder_info["relative_folder"], folder_info["path"], folder_info["file_count"], extension_text])
        if not workbook.worksheets:
            continue
        output = _unique_export_path(folder)
        workbook.save(output)
        outputs.append(output)
        generated_tables[folder_text] = output
    # Generate the index after the workbooks so a failed workbook write cannot
    # leave a misleading map behind.
    if by_folder:
        filetree = _write_filetree_export(root, generated_tables, include_hidden, include_system)
        if filetree is not None:
            outputs.append(filetree)
    return outputs


def open_in_explorer(path: str | Path) -> None:
    path = str(Path(path).resolve())
    if os.name == "nt":
        subprocess.Popen(["explorer", path])
    elif os.name == "darwin":
        subprocess.Popen(["open", path])
    else:
        subprocess.Popen(["xdg-open", path])
