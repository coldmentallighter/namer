"""XLSX import/export, directory statistics and filetree exports.

Split out of the former ``core/files.py`` god module.  This is the only core
module that touches openpyxl.
"""

from __future__ import annotations

import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterable, Sequence

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - UI reports this clearly
    Workbook = None
    load_workbook = None

from .fsutil import is_hidden, is_system, natural_key, normalise_ext
from .models import ExcelMatchResult, FileRecord, NamingGroup
from .scan import scan_folder


def _excel_value(value) -> str:
    return "" if value is None else str(value).strip()


def _match_key(value: str, extension: str = "") -> str:
    value = _excel_value(value)
    if not value:
        return ""
    # Accept a bare filename or a Windows/POSIX path pasted into Excel.
    value = re.split(r"[\\/]", value)[-1]
    suffix = normalise_ext(extension)
    if suffix and value.casefold().endswith(suffix.casefold()):
        value = value[: -len(suffix)]
    return value.casefold()


def _excel_name(value: str, extension: str) -> str:
    value = _excel_value(value)
    if not value:
        return ""
    suffix = normalise_ext(extension)
    if suffix and value.casefold().endswith(suffix.casefold()):
        return value[: -len(suffix)]
    return value


_EXCEL_NAME_TEMPLATE = re.compile(r"\{([a-zA-Z][a-zA-Z0-9_.-]*)\}")


def _expand_excel_name_template(value: str, record: FileRecord,
                                row_values: dict[str, str] | None = None,
                                value_expander: Callable[[str, FileRecord, dict[str, str]], str] | None = None) -> str:
    """Expand placeholders through a caller-supplied workflow policy.

    The core fallback only uses values from the spreadsheet header row. A
    workflow can provide ``value_expander`` to combine those values with its
    own metadata namespace and normalizers.
    """
    source = _excel_name(value, record.extension)
    values = {str(key).casefold(): _excel_value(item)
              for key, item in (row_values or {}).items()}
    if value_expander:
        return value_expander(source, record, values)
    if not _EXCEL_NAME_TEMPLATE.search(source):
        return source
    expanded = _EXCEL_NAME_TEMPLATE.sub(
        lambda match: values.get(match.group(1).casefold(), ""), source
    )
    expanded = re.sub(r"([ _.-])\1+", r"\1", expanded)
    return expanded.strip(" _-.")


def import_xlsx(xlsx_path: str | Path, group: NamingGroup,
                sheet_name: str | None = None,
                value_expander: Callable[[str, FileRecord, dict[str, str]], str] | None = None) -> ExcelMatchResult:
    if load_workbook is None:
        raise RuntimeError("需要安装 openpyxl 才能导入 XLSX")
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet = None
        if sheet_name:
            sheet = next((candidate for candidate in workbook.worksheets
                          if candidate.title.casefold() == str(sheet_name).casefold()), None)
        if sheet is None:
            wanted = group.extension.lstrip(".").casefold()
            sheet = next((candidate for candidate in workbook.worksheets
                          if candidate.title.casefold() == wanted), workbook.worksheets[0])
        selected_sheet_name = sheet.title
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        # read_only workbooks keep a zip handle open on Windows until explicitly closed.
        workbook.close()
    mapping: dict[str, str] = {}
    unmatched_rows: list[tuple[int, str, str]] = []
    warnings: list[str] = []
    matched_without_name: list[FileRecord] = []
    records_by_key = {_match_key(record.original_name, group.extension): record for record in group.records}
    unmatched_files = list(group.records)
    detail_mode = False
    headers: list[str] = []
    if rows:
        headers = [str(value or "").strip().casefold() for value in rows[0]]
        header_tokens = {"source", "sourcename", "source name", "原文件名", "源文件名", "newname", "new name", "新名称", "新文件名", "relativepath", "relative path", "相对路径"}
        if any(token in header_tokens for token in headers):
            detail_mode = True
            source_index = next((index for index, value in enumerate(headers)
                                 if value in {"source", "sourcename", "source name", "原文件名", "源文件名"}), 0)
            name_index = next((index for index, value in enumerate(headers)
                               if value in {"newname", "new name", "新名称", "新文件名", "name", "名称"}), 1 if len(headers) > 1 else None)
            data_rows = rows[1:]
        else:
            source_index, name_index, data_rows = 0, 1 if sheet.max_column >= 2 else None, rows
    else:
        source_index, name_index, data_rows = 0, None, []
    has_b = name_index is not None and (detail_mode or sheet.max_column >= 2)
    if has_b:
        mode = "source-and-name"
        for row_number, row in enumerate(data_rows, start=2 if detail_mode else 1):
            source = _excel_value(row[source_index] if len(row) > source_index else "")
            new_name = _excel_value(row[name_index] if name_index is not None and len(row) > name_index else "")
            if not source and not new_name:
                continue
            key = _match_key(source, group.extension)
            record = records_by_key.get(key)
            if not record:
                unmatched_rows.append((row_number, source, new_name))
                warnings.append(f"WARN Excel 第 {row_number} 行无法匹配文件: {source}")
                continue
            if not new_name:
                warnings.append(f"WARN Excel 第 {row_number} 行 B 列为空，跳过: {source}")
                if record in unmatched_files:
                    unmatched_files.remove(record)
                matched_without_name.append(record)
                continue
            row_values = {
                headers[index]: _excel_value(cell)
                for index, cell in enumerate(row)
                if index < len(headers) and headers[index]
            }
            new_name = _expand_excel_name_template(new_name, record, row_values, value_expander)
            if not new_name:
                warnings.append(f"WARN Excel 第 {row_number} 行 B 列模板展开后为空，跳过: {source}")
                if record in unmatched_files:
                    unmatched_files.remove(record)
                matched_without_name.append(record)
                continue
            if record.path in mapping:
                warnings.append(f"WARN Excel 第 {row_number} 行重复匹配: {source}")
                continue
            mapping[record.path] = new_name
            record.excel_source = source
            if record in unmatched_files:
                unmatched_files.remove(record)
    else:
        mode = "ordered-names"
        row_index = 0
        for row_number, row in enumerate(data_rows, start=2 if detail_mode else 1):
            value = _excel_value(row[0] if row else "")
            if not value:
                unmatched_rows.append((row_number, "", ""))
                warnings.append(f"WARN Excel 第 {row_number} 行为空")
                row_index += 1
                continue
            if row_index >= len(group.records):
                unmatched_rows.append((row_number, value, ""))
                warnings.append(f"WARN Excel 第 {row_number} 行超出当前命名组文件数")
                row_index += 1
                continue
            record = group.records[row_index]
            mapping[record.path] = _excel_name(value, group.extension)
            record.excel_source = value
            if record in unmatched_files:
                unmatched_files.remove(record)
            row_index += 1
        if len(group.records) > row_index:
            warnings.append(f"WARN {len(group.records) - row_index} 个文件未分配 Excel 名称")
    return ExcelMatchResult(mode, mapping, len(mapping), unmatched_files, unmatched_rows, warnings,
                            matched_without_name, selected_sheet_name, detail_mode)


def _unique_export_path(folder: Path) -> Path:
    base = folder.name or "root"
    candidate = folder.parent / f"{base}.ffnf.xlsx"
    if not candidate.exists():
        return candidate
    index = 1
    while True:
        candidate = folder.parent / f"{base}.ori{index:02d}.ffnf.xlsx"
        if not candidate.exists():
            return candidate
        index += 1


def _structure_directories(root: str | Path, include_hidden: bool = False,
                           include_system: bool = False) -> tuple[dict[Path, list[Path]], int]:
    """Collect a filtered directory tree and return its maximum relative depth."""
    root_path = Path(root).expanduser().resolve()
    children: dict[Path, list[Path]] = {}
    max_depth = 0
    for directory, dirs, _files in os.walk(root_path):
        directory_path = Path(directory)
        dirs[:] = sorted(dirs, key=natural_key)
        if not include_hidden:
            dirs[:] = [name for name in dirs if not is_hidden(directory_path / name)]
        if not include_system:
            dirs[:] = [name for name in dirs if not is_system(directory_path / name)]
        child_paths = [directory_path / name for name in dirs]
        children[directory_path] = child_paths
        for child in child_paths:
            try:
                depth = len(child.relative_to(root_path).parts)
            except ValueError:
                continue
            max_depth = max(max_depth, depth)
    children.setdefault(root_path, [])
    return children, max_depth


def _write_filetree_export(root: str | Path, generated_tables: dict[str, Path],
                           include_hidden: bool = False,
                           include_system: bool = False) -> Path | None:
    """Write an index of content directories that received an XLSX export."""
    root_path = Path(root).expanduser().resolve()
    _children, max_depth = _structure_directories(root_path, include_hidden, include_system)
    if max_depth < 3 or not generated_tables:
        return None
    output = root_path / "filetree.txt"
    if output.exists():
        return None

    lines = [
        f"根目录: {root_path}",
        "生成 .ffnf.xlsx 的内容目录:",
    ]
    ordered = sorted(
        generated_tables.items(),
        key=lambda pair: natural_key(os.path.relpath(pair[0], root_path)),
    )
    for folder_text, workbook in ordered:
        folder = Path(folder_text).resolve()
        try:
            relative_folder = folder.relative_to(root_path).as_posix() or "."
        except ValueError:
            relative_folder = os.path.relpath(folder, root_path).replace(os.sep, "/")
        relative_workbook = os.path.relpath(Path(workbook).resolve(), root_path).replace(os.sep, "/")
        lines.append(f"- {relative_folder}")
        lines.append(f"  XLSX: {relative_workbook}")
    output.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return output


def collect_directory_statistics(root: str | Path, records: Sequence[FileRecord] | None = None,
                                 include_hidden: bool = False, include_system: bool = False) -> dict[str, Any]:
    root_path = Path(root).expanduser().resolve()
    result = scan_folder(root_path, include_hidden, include_system) if records is None else None
    effective_records = list(records if records is not None else result.records)
    tree, _max_depth = _structure_directories(root_path, include_hidden, include_system)
    directories: set[str] = {str(path) for path in tree}
    per_folder: dict[str, dict[str, Any]] = {}
    extension_counts: dict[str, int] = {}
    for record in effective_records:
        folder = str(Path(record.path).parent)
        directories.add(folder)
        info = per_folder.setdefault(folder, {"path": folder, "relative_folder": record.relative_folder,
                                               "file_count": 0, "extensions": {}})
        info["file_count"] += 1
        ext = record.extension or "(none)"
        info["extensions"][ext] = info["extensions"].get(ext, 0) + 1
        extension_counts[ext] = extension_counts.get(ext, 0) + 1
    return {
        "root": str(root_path),
        "directory_count": len(directories),
        "file_count": len(effective_records),
        "content_directory_count": len(per_folder),
        "empty_directory_count": max(0, len(directories) - len(per_folder)),
        "extension_counts": dict(sorted(extension_counts.items(), key=lambda pair: natural_key(pair[0]))),
        "folders": sorted(per_folder.values(), key=lambda item: natural_key(item["path"])),
    }


def _flatten_metadata(value: Any, prefix: str = "") -> dict[str, Any]:
    flattened: dict[str, Any] = {}
    if isinstance(value, dict):
        for key, child in value.items():
            child_prefix = f"{prefix}.{key}" if prefix else str(key)
            flattened.update(_flatten_metadata(child, child_prefix))
    elif prefix:
        flattened[prefix] = value
    return flattened


def export_filename_tables(root: str | Path, selected_extensions: Iterable[str],
                           include_hidden: bool = False, include_system: bool = False,
                           metadata_reader: Callable[[str | Path, str | Path | None], dict[str, Any]] | None = None) -> list[Path]:
    if Workbook is None:
        raise RuntimeError("需要安装 openpyxl 才能导出 XLSX")
    selected = {normalise_ext(ext).casefold() for ext in selected_extensions}
    result = scan_folder(root, include_hidden=include_hidden, include_system=include_system,
                         metadata_reader=metadata_reader)
    by_folder: dict[str, list[FileRecord]] = {}
    for record in result.records:
        if record.extension.casefold() in selected:
            by_folder.setdefault(record.path and str(Path(record.path).parent), []).append(record)
    outputs: list[Path] = []
    generated_tables: dict[str, Path] = {}
    for folder_text, records in sorted(by_folder.items(), key=lambda pair: natural_key(pair[0])):
        folder = Path(folder_text)
        workbook = Workbook()
        workbook.remove(workbook.active)
        by_ext: dict[str, list[FileRecord]] = {}
        for record in records:
            by_ext.setdefault(record.extension.lower(), []).append(record)
        used_titles: set[str] = set()
        for ext, ext_records in sorted(by_ext.items(), key=lambda pair: natural_key(pair[0])):
            title = ext.lstrip(".").upper()[:31] or "NO_EXT"
            original_title = title
            suffix = 1
            while title in used_titles:
                title = f"{original_title[:28]}_{suffix}"
                suffix += 1
            used_titles.add(title)
            sheet = workbook.create_sheet(title)
            ordered_records = sorted(ext_records, key=lambda item: natural_key(item.stem))
            metadata_values = [_flatten_metadata(record.metadata) for record in ordered_records]
            metadata_keys = sorted({key for values in metadata_values for key in values}, key=natural_key)
            metadata_headers = [f"Metadata.{key}" for key in metadata_keys]
            sheet.append(["SourceName", "NewName", "RelativePath", "Folder", "Extension", "SizeBytes", "ModifiedTime", "Association", *metadata_headers])
            for record in ordered_records:
                try:
                    stat = Path(record.path).stat()
                    size, modified = int(stat.st_size), datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds")
                except OSError:
                    size, modified = "", ""
                values = _flatten_metadata(record.metadata)
                sheet.append([record.stem, "", record.relative_folder, record.folder_name,
                              record.extension, size, modified, record.association_id,
                              *[values.get(key, "") for key in metadata_keys]])
        stats = collect_directory_statistics(root, result.records, include_hidden, include_system)
        metadata = workbook.create_sheet("Metadata")
        metadata.append(["Field", "Value"])
        metadata.append(["Root", stats["root"]])
        metadata.append(["DirectoryCount", stats["directory_count"]])
        metadata.append(["FileCount", stats["file_count"]])
        metadata.append(["ContentDirectoryCount", stats["content_directory_count"]])
        metadata.append(["AssociationCount", len(result.associations)])
        summary = workbook.create_sheet("Summary")
        summary.append(["RelativeFolder", "Path", "FileCount", "Extensions"])
        for folder_info in stats["folders"]:
            extension_text = ", ".join(f"{key}:{value}" for key, value in folder_info["extensions"].items())
            summary.append([folder_info["relative_folder"], folder_info["path"], folder_info["file_count"], extension_text])
        if not workbook.worksheets:
            continue
        output = _unique_export_path(folder)
        workbook.save(output)
        outputs.append(output)
        generated_tables[folder_text] = output
    # Generate the index after the workbooks so a failed workbook write cannot
    # leave a misleading map behind.
    if by_folder:
        filetree = _write_filetree_export(root, generated_tables, include_hidden, include_system)
        if filetree is not None:
            outputs.append(filetree)
    return outputs
