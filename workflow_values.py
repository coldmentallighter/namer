"""Persistent XLSX vocabularies used by the workflow tag manager.

Workflow structure remains declarative JSON. This module stores the editable
values for each workflow in one workbook, with one worksheet per field.
"""

from __future__ import annotations

import copy
import os
import re
import threading
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:  # pragma: no cover - the application dependency is pinned.
    Workbook = None
    load_workbook = None


WORKBOOK_EXTENSION = ".xlsx"
WORKBOOK_VERSION = 1
TAG_HEADERS = ["TagId", "显示名称", "实际值", "别名", "启用", "默认", "排序", "使用次数", "来源", "备注"]
_SAFE_ID = re.compile(r"^[a-z][a-z0-9_-]{0,63}$")
_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def _text(value: Any) -> str:
    return str(value or "").strip()


def _bool(value: Any, default: bool = True) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    return _text(value).casefold() in {"1", "true", "yes", "y", "是", "启用", "已启用"}


def _int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _aliases(value: Any) -> list[str]:
    if isinstance(value, list):
        values = value
    else:
        values = re.split(r"[，,、]", _text(value))
    result: list[str] = []
    for item in values:
        item = _text(item)
        if item and item not in result:
            result.append(item)
    return result


def _normalise_tag(value: Any, index: int = 0, source: str = "工作流默认") -> dict[str, Any]:
    raw = {"label": value, "value": value} if isinstance(value, str) else (value if isinstance(value, dict) else {})
    label = _text(raw.get("label") or raw.get("value"))
    actual = _text(raw.get("value") or raw.get("label"))
    tag_id = _text(raw.get("id"))
    if not tag_id:
        slug = re.sub(r"[^a-z0-9]+", "-", actual.casefold()).strip("-") or "value"
        tag_id = f"tag-{index + 1}-{slug}"
    return {
        "id": tag_id,
        "label": label,
        "value": actual,
        "aliases": _aliases(raw.get("aliases", [])),
        "enabled": _bool(raw.get("enabled"), True),
        "default": _bool(raw.get("default"), False),
        "order": _int(raw.get("order"), index + 1),
        "usage": _int(raw.get("usage"), 0),
        "source": _text(raw.get("source")) or source,
        "notes": _text(raw.get("notes")),
    }


class WorkflowValueStore:
    """Read and atomically write one editable workbook per workflow."""

    def __init__(self, root: str | Path) -> None:
        self.root = Path(root)
        self.lock = threading.RLock()

    def path_for(self, workflow_id: str) -> Path:
        workflow_id = _text(workflow_id)
        if not _SAFE_ID.fullmatch(workflow_id):
            raise ValueError("工作流 id 无效")
        return self.root / f"{workflow_id}.xlsx"

    def _sheet_names(self, workflow: dict[str, Any]) -> dict[str, str]:
        used: set[str] = {"00_说明"}
        result: dict[str, str] = {}
        for index, field in enumerate(workflow.get("fields", []), start=1):
            field_id = _text(field.get("id"))
            label = _text(field.get("label")) or field_id
            base = _INVALID_SHEET_CHARS.sub(" ", label).strip() or field_id
            base = re.sub(r"\s+", "_", base)
            prefix = f"{index:02d}_"
            name = (prefix + base)[:31]
            if name in used:
                suffix = 2
                while True:
                    candidate = f"{name[:27]}_{suffix}"
                    if candidate not in used:
                        name = candidate
                        break
                    suffix += 1
            used.add(name)
            result[field_id] = name
        return result

    def _seed_tags(self, workflow: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
        return {
            _text(field.get("id")): [
                _normalise_tag(tag, index, "工作流默认")
                for index, tag in enumerate(field.get("quick_tags", []))
            ]
            for field in workflow.get("fields", [])
        }

    def _read_sheet(self, sheet: Any) -> list[dict[str, Any]]:
        header_row = None
        header_index: dict[str, int] = {}
        for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            values = [_text(value) for value in row]
            if "TagId" not in values:
                continue
            header_row = row_number
            header_index = {value: index for index, value in enumerate(values) if value}
            break
        if header_row is None:
            return []
        result: list[dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            def cell(name: str) -> Any:
                index = header_index.get(name)
                return row[index] if index is not None and index < len(row) else ""

            tag_id = _text(cell("TagId"))
            label = _text(cell("显示名称"))
            value = _text(cell("实际值"))
            if not tag_id and not label and not value:
                continue
            result.append({
                "id": tag_id,
                "label": label,
                "value": value,
                "aliases": _aliases(cell("别名")),
                "enabled": _bool(cell("启用"), True),
                "default": _bool(cell("默认"), False),
                "order": _int(cell("排序"), len(result) + 1),
                "usage": _int(cell("使用次数"), 0),
                "source": _text(cell("来源")) or "XLSX",
                "notes": _text(cell("备注")),
            })
        return result

    def read(self, workflow: dict[str, Any]) -> dict[str, Any]:
        if Workbook is None or load_workbook is None:
            raise RuntimeError("需要安装 openpyxl 才能使用工作流标签 XLSX")
        path = self.path_for(_text(workflow.get("id")))
        seeded = self._seed_tags(workflow)
        sheets = self._sheet_names(workflow)
        exists = path.is_file()
        tags = copy.deepcopy(seeded)
        with self.lock:
            if exists:
                try:
                    workbook = load_workbook(path, read_only=True, data_only=True)
                except Exception as exc:
                    raise ValueError(f"工作流值 XLSX 无法读取: {path.name}: {exc}") from exc
                try:
                    for field_id, sheet_name in sheets.items():
                        if sheet_name in workbook.sheetnames:
                            tags[field_id] = self._read_sheet(workbook[sheet_name])
                finally:
                    workbook.close()
        return {
            "workflow_id": _text(workflow.get("id")),
            "workbook": str(path),
            "workbook_name": path.name,
            "exists": exists,
            "version": WORKBOOK_VERSION,
            "tags": tags,
            "sheets": [
                {"field_id": field_id, "name": sheet_name, "label": _text(field.get("label"))}
                for field_id, sheet_name in sheets.items()
                for field in workflow.get("fields", [])
                if _text(field.get("id")) == field_id
            ],
        }

    @staticmethod
    def _style_sheet(sheet: Any) -> None:
        accent = "16715D"
        light = "DFF1EA"
        line = Side(style="thin", color="DCE5E1")
        sheet.freeze_panes = "A5"
        sheet.auto_filter.ref = f"A4:J{max(4, sheet.max_row)}"
        for cell in sheet[1]:
            cell.font = Font(size=14, bold=True, color=accent)
        for cell in sheet[4]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=accent)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = Border(bottom=line)
        for row in sheet.iter_rows(min_row=5):
            for cell in row:
                cell.border = Border(bottom=line)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        widths = {"A": 24, "B": 20, "C": 22, "D": 28, "E": 10, "F": 10, "G": 10, "H": 12, "I": 16, "J": 28}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        sheet.sheet_view.showGridLines = False
        sheet[2][0].fill = PatternFill("solid", fgColor=light)
        sheet[2][0].font = Font(color=accent, italic=True)

    def write(self, workflow: dict[str, Any], tags: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
        if Workbook is None:
            raise RuntimeError("需要安装 openpyxl 才能写入工作流标签 XLSX")
        workflow_id = _text(workflow.get("id"))
        path = self.path_for(workflow_id)
        sheet_names = self._sheet_names(workflow)
        workbook = Workbook()
        workbook.remove(workbook.active)
        overview = workbook.create_sheet("00_说明")
        overview.append([f"工作流值 · {_text(workflow.get('name'))}"])
        overview.append(["WorkbookVersion", WORKBOOK_VERSION])
        overview.append(["WorkflowId", workflow_id])
        overview.append([])
        overview.append(["FieldId", "字段名称", "作用域", "类型", "工作表", "模板顺序"])
        template_order = [part.get("field") for part in workflow.get("template", []) if part.get("field")]
        for field in workflow.get("fields", []):
            field_id = _text(field.get("id"))
            overview.append([
                field_id,
                _text(field.get("label")) or field_id,
                _text(field.get("scope")),
                _text(field.get("kind")),
                sheet_names[field_id],
                template_order.index(field_id) + 1 if field_id in template_order else "",
            ])
        overview.freeze_panes = "A6"
        overview.auto_filter.ref = f"A5:F{max(5, overview.max_row)}"
        overview.sheet_view.showGridLines = False
        for cell in overview[1]:
            cell.font = Font(size=14, bold=True, color="16715D")
        for cell in overview[5]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="16715D")
        for column, width in {"A": 24, "B": 22, "C": 14, "D": 14, "E": 28, "F": 14}.items():
            overview.column_dimensions[column].width = width

        for field in workflow.get("fields", []):
            field_id = _text(field.get("id"))
            sheet = workbook.create_sheet(sheet_names[field_id])
            label = _text(field.get("label")) or field_id
            sheet.append([f"快捷标签 · {label}"])
            sheet.append([f"字段 ID：{field_id} · 作用域：{_text(field.get('scope'))}"])
            sheet.append([])
            sheet.append(TAG_HEADERS)
            values = sorted(tags.get(field_id, []), key=lambda item: (_int(item.get("order"), 999999), _text(item.get("label")).casefold()))
            for index, raw in enumerate(values, start=1):
                tag = _normalise_tag(raw, index - 1, _text(raw.get("source")) or "用户新增")
                sheet.append([
                    tag["id"], tag["label"], tag["value"], "、".join(tag["aliases"]),
                    "是" if tag["enabled"] else "否", "是" if tag["default"] else "否",
                    tag["order"], tag["usage"], tag["source"], tag["notes"],
                ])
            self._style_sheet(sheet)

        self.root.mkdir(parents=True, exist_ok=True)
        temporary = path.with_suffix(path.suffix + ".tmp")
        try:
            with self.lock:
                workbook.save(temporary)
                os.replace(temporary, path)
        except OSError as exc:
            try:
                temporary.unlink(missing_ok=True)
            except OSError:
                pass
            raise ValueError(f"无法保存工作流值 XLSX，请确认文件未被 Excel 占用: {exc}") from exc
        return self.read(workflow)

    def upsert(self, workflow: dict[str, Any], field_id: str, value: dict[str, Any]) -> dict[str, Any]:
        field = next((item for item in workflow.get("fields", []) if _text(item.get("id")) == field_id), None)
        if field is None:
            raise ValueError(f"工作流字段不存在: {field_id}")
        if _text(field.get("kind")) == "fixed":
            raise ValueError("固定字段不能添加快捷标签")
        payload = self.read(workflow)
        current = payload["tags"].setdefault(field_id, [])
        tag = _normalise_tag(value, len(current), "用户新增")
        if not tag["label"] or not tag["value"]:
            raise ValueError("标签显示名称和实际值不能为空")
        tag_id = _text(value.get("id"))
        duplicate = next((item for item in current if item["id"] != tag_id and item["value"].casefold() == tag["value"].casefold()), None)
        if duplicate:
            raise ValueError(f"实际值已存在: {duplicate['label']}")
        if tag_id:
            existing = next((item for item in current if item["id"] == tag_id), None)
            if existing:
                tag["usage"] = existing.get("usage", 0)
                tag["source"] = existing.get("source") or "用户编辑"
                current[current.index(existing)] = tag
            else:
                current.append(tag)
        else:
            tag["id"] = f"tag-{field_id}-{len(current) + 1}"
            while any(item["id"] == tag["id"] for item in current):
                tag["id"] += "-2"
            tag["order"] = max([_int(item.get("order"), 0) for item in current] or [0]) + 1
            current.append(tag)
        return self.write(workflow, payload["tags"])

    def toggle(self, workflow: dict[str, Any], field_id: str, tag_id: str) -> dict[str, Any]:
        payload = self.read(workflow)
        current = payload["tags"].get(field_id, [])
        tag = next((item for item in current if item["id"] == tag_id), None)
        if tag is None:
            raise ValueError("标签不存在")
        tag["enabled"] = not bool(tag.get("enabled", True))
        return self.write(workflow, payload["tags"])
