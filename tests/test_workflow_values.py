from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from workflow_system.values import WorkflowValueStore


class WorkflowValueStoreTests(unittest.TestCase):
    WORKFLOW = {
        "id": "tag-test",
        "name": "标签测试工作流",
        "template": [{"field": "meta_prefix"}, {"field": "name"}],
        "fields": [
            {
                "id": "meta_prefix",
                "label": "元前缀",
                "scope": "group",
                "kind": "text",
                "quick_tags": [{"label": "CL_", "value": "CL_"}],
            },
            {"id": "name", "label": "名称", "scope": "record", "kind": "text"},
        ],
    }

    def test_seed_write_reload_toggle_and_workbook_shape(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = WorkflowValueStore(Path(temp_dir))

            initial = store.read(self.WORKFLOW)
            self.assertFalse(initial["exists"])
            self.assertEqual(initial["tags"]["meta_prefix"][0]["value"], "CL_")
            self.assertEqual(initial["tags"]["name"], [])

            saved = store.upsert(self.WORKFLOW, "meta_prefix", {
                "label": "Cymatics",
                "value": "Cymatics_",
                "aliases": "Cymatics, Cym",
                "source": "解析确认",
            })
            workbook_path = Path(saved["workbook"])
            self.assertTrue(saved["exists"])
            self.assertTrue(workbook_path.is_file())
            self.assertEqual(saved["workbook_name"], "tag-test.xlsx")

            reloaded = store.read(self.WORKFLOW)
            tags = reloaded["tags"]["meta_prefix"]
            cymatics = next(tag for tag in tags if tag["value"] == "Cymatics_")
            self.assertEqual(cymatics["aliases"], ["Cymatics", "Cym"])
            self.assertTrue(cymatics["enabled"])

            toggled = store.toggle(self.WORKFLOW, "meta_prefix", cymatics["id"])
            disabled = next(tag for tag in toggled["tags"]["meta_prefix"] if tag["id"] == cymatics["id"])
            self.assertFalse(disabled["enabled"])
            self.assertFalse(next(tag for tag in store.read(self.WORKFLOW)["tags"]["meta_prefix"] if tag["id"] == cymatics["id"])["enabled"])

            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["00_说明", "01_元前缀", "02_名称"])
                sheet = workbook["01_元前缀"]
                self.assertEqual([cell.value for cell in sheet[4]], [
                    "TagId", "显示名称", "实际值", "别名", "启用", "默认", "排序", "使用次数", "来源", "备注",
                ])
                row = next(row for row in sheet.iter_rows(min_row=5, values_only=True) if row[2] == "Cymatics_")
                self.assertEqual(row[1], "Cymatics")
                self.assertEqual(row[3], "Cymatics、Cym")
                self.assertEqual(row[4], "否")
                self.assertEqual(row[8], "解析确认")
            finally:
                workbook.close()

            deleted = store.delete(self.WORKFLOW, "meta_prefix", cymatics["id"])
            self.assertNotIn(cymatics["id"], [tag["id"] for tag in deleted["tags"]["meta_prefix"]])
            self.assertNotIn(cymatics["id"], [tag["id"] for tag in store.read(self.WORKFLOW)["tags"]["meta_prefix"]])

    def test_duplicate_values_and_fixed_fields_are_rejected(self):
        workflow = {
            **self.WORKFLOW,
            "fields": [*self.WORKFLOW["fields"], {"id": "fixed", "label": "固定", "kind": "fixed"}],
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            store = WorkflowValueStore(Path(temp_dir))
            store.upsert(workflow, "meta_prefix", {"label": "One", "value": "one"})
            with self.assertRaisesRegex(ValueError, "实际值已存在"):
                store.upsert(workflow, "meta_prefix", {"label": "Duplicate", "value": "ONE"})
            with self.assertRaisesRegex(ValueError, "固定字段"):
                store.upsert(workflow, "fixed", {"label": "Nope", "value": "nope"})
            with self.assertRaisesRegex(ValueError, "标签不存在"):
                store.delete(workflow, "meta_prefix", "missing-tag")


if __name__ == "__main__":
    unittest.main()
