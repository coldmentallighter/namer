from __future__ import annotations

import json
import struct
import tempfile
import unittest
import wave
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from namer_core import (
    NamingGroup,
    assign_numeric,
    build_stem_associations,
    collect_directory_statistics,
    compose_filename,
    directory_prefix_defaults,
    execute_rename,
    export_filename_tables,
    file_fingerprint,
    import_xlsx,
    read_file_metadata,
    natural_key,
    parse_filename,
    preflight,
    redo_last,
    scan_folder,
    undo_last,
    validate_filename,
)
from workflow_metadata import parse_workflow_filename, read_workflow_metadata
from workflow_config import CORE_FALLBACK_WORKFLOW, discover_workflows, validate_workflow
from workflow_runtime import WorkflowModuleRegistry


INSTALLED_WORKFLOWS, _WORKFLOW_ERRORS = discover_workflows()
WORKFLOW_MODULE_REGISTRY = WorkflowModuleRegistry()
_WORKFLOW_MODULE_ERRORS = WORKFLOW_MODULE_REGISTRY.refresh(
    {
        workflow_id: workflow["_source_dir"]
        for workflow_id, workflow in INSTALLED_WORKFLOWS.items()
        if workflow.get("_source_dir")
    },
    INSTALLED_WORKFLOWS,
)
if _WORKFLOW_MODULE_ERRORS:
    raise RuntimeError(f"测试工作流模块加载失败: {_WORKFLOW_MODULE_ERRORS}")
SAMPLE_PACK_MODULE = WORKFLOW_MODULE_REGISTRY.module("sample-pack", "sample_pack")
IMAGE_ASSETS_MODULE = WORKFLOW_MODULE_REGISTRY.module("wallpaper-assets", "image_assets")
append_bpm_suffix = SAMPLE_PACK_MODULE.append_bpm_suffix
detect_bpm = SAMPLE_PACK_MODULE.detect_bpm
read_image_dimensions = IMAGE_ASSETS_MODULE.read_image_dimensions
DEFAULT_TEST_WORKFLOW = INSTALLED_WORKFLOWS.get("default", CORE_FALLBACK_WORKFLOW)
SAMPLE_PACK_WORKFLOW = INSTALLED_WORKFLOWS.get("sample-pack")
IMAGE_METADATA_WORKFLOW = validate_workflow({
    **INSTALLED_WORKFLOWS["wallpaper-assets"],
})
SAMPLE_METADATA_WORKFLOW = validate_workflow({**INSTALLED_WORKFLOWS["sample-pack"]})


class NamerCoreTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name) / "根目录 [EDM]"
        (self.root / "Drums&Loop").mkdir(parents=True)
        (self.root / "MIDI").mkdir()
        (self.root / "Drums&Loop" / "Kick 2.wav").write_bytes(b"RIFF")
        (self.root / "Drums&Loop" / "Kick 10.WAV").write_bytes(b"RIFF")
        (self.root / "Drums&Loop" / "中文 #&-_.png").write_bytes(b"PNG")
        (self.root / "MIDI" / "Pattern1.mid").write_bytes(b"MThd")
        (self.root / "MIDI" / "README").write_text("plain", encoding="utf-8")
        (self.root / "MIDI" / ".hidden.wav").write_bytes(b"RIFF")
        (self.root / "Drums&Loop" / "existing.wav").write_bytes(b"x")
        # This workbook must never be counted or exported again.
        (self.root / "Drums&Loop.ffnf.xlsx").write_bytes(b"generated")

    def tearDown(self):
        self.tmp.cleanup()

    @staticmethod
    def _write_bpm_wav(path: Path, bpm: int) -> None:
        value = f"{bpm}\0".encode("ascii")
        value += b"\0" * (len(value) & 1)
        info = b"INFO" + b"TBPM" + struct.pack("<I", len(value)) + value
        fmt = struct.pack("<HHIIHH", 1, 1, 8_000, 16_000, 2, 16)
        body = (
            b"WAVE"
            + b"fmt " + struct.pack("<I", len(fmt)) + fmt
            + b"LIST" + struct.pack("<I", len(info)) + info
            + b"data" + struct.pack("<I", 0)
        )
        path.write_bytes(b"RIFF" + struct.pack("<I", len(body)) + body)

    @staticmethod
    def _write_bpm_midi(path: Path, bpm: int) -> None:
        microseconds = round(60_000_000 / bpm)
        track = b"\x00\xff\x51\x03" + microseconds.to_bytes(3, "big") + b"\x00\xff\x2f\x00"
        path.write_bytes(
            b"MThd" + struct.pack(">IHHH", 6, 0, 1, 480)
            + b"MTrk" + struct.pack(">I", len(track)) + track
        )

    def test_scan_dynamic_extensions_groups_and_hidden(self):
        result = scan_folder(self.root)
        self.assertEqual(result.extension_counts[".wav"], 3)
        self.assertEqual(result.extension_counts[".png"], 1)
        self.assertEqual(result.extension_counts[".mid"], 1)
        self.assertEqual(result.extension_counts[""], 1)
        self.assertNotIn(".xlsx", result.extension_counts)
        self.assertTrue(any("Drums&Loop" in key for key in result.groups))
        shown = {Path(record.path).name for record in result.records}
        self.assertNotIn(".hidden.wav", shown)
        result_with_hidden = scan_folder(self.root, include_hidden=True)
        self.assertIn(".hidden.wav", {Path(record.path).name for record in result_with_hidden.records})

    def test_natural_sort_and_name_rules(self):
        self.assertEqual(sorted(["Kick 10.wav", "Kick 2.wav"], key=natural_key), ["Kick 2.wav", "Kick 10.wav"])
        self.assertEqual(compose_filename("[EDM] 根", "Drums&Loop", "Kick", "Chest", ".wav"), "[EDM] 根_Drums&Loop_Kick_Chest.wav")
        self.assertEqual(compose_filename("  Root  ", "Folder", "", "Name", ".wav"), "  Root  _Folder_Name.wav")
        self.assertEqual(compose_filename("", "", "", "Chest", ".WAV"), "Chest.WAV")
        self.assertIn("非法", validate_filename("a:b.wav"))
        self.assertIn("保留设备名", validate_filename("CON.txt.wav"))
        self.assertIsNone(validate_filename("中文 #&-_ sample.wav"))

    def test_common_metadata_does_not_classify_format_specific_namespaces(self):
        metadata = read_file_metadata(self.root / "Drums&Loop" / "Kick 2.wav", self.root)
        self.assertTrue(metadata["file"]["mime_type"].startswith("audio/"))
        self.assertNotIn("image", metadata)
        signature_image = self.root / "Drums&Loop" / "signature.dat"
        signature_image.write_bytes(
            b"\x89PNG\r\n\x1a\n" + (13).to_bytes(4, "big") + b"IHDR"
            + (640).to_bytes(4, "big") + (480).to_bytes(4, "big")
            + b"\x08\x02\x00\x00\x00"
        )
        self.assertNotIn("image", read_workflow_metadata(
            DEFAULT_TEST_WORKFLOW, signature_image, self.root
        ))
        self.assertNotIn("sample_pack", read_workflow_metadata(
            SAMPLE_METADATA_WORKFLOW, signature_image, self.root
        ))
        image_metadata = read_workflow_metadata(
            IMAGE_METADATA_WORKFLOW, signature_image, self.root
        )
        self.assertEqual(image_metadata["image"]["aspect_ratio"], "4:3")
        unknown_sample = self.root / "Drums&Loop" / "Loop_128.bin"
        unknown_sample.write_bytes(b"not a format-specific header")
        sample_metadata = read_workflow_metadata(
            SAMPLE_METADATA_WORKFLOW, unknown_sample, self.root
        )
        self.assertEqual(sample_metadata["sample_pack"]["bpm"], "128")
        self.assertNotIn("sample_pack", read_workflow_metadata(
            DEFAULT_TEST_WORKFLOW, unknown_sample, self.root
        ))
        image_metadata = read_workflow_metadata(
            IMAGE_METADATA_WORKFLOW, self.root / "Drums&Loop" / "中文 #&-_.png", self.root
        )
        self.assertNotIn("image", image_metadata)

    def test_numeric_is_independent_per_group(self):
        result = scan_folder(self.root)
        wav_group = next(group for group in result.groups.values() if group.extension == ".wav" and group.folder_name == "Drums&Loop")
        mid_group = next(group for group in result.groups.values() if group.extension == ".mid")
        assign_numeric(wav_group, meta_prefix="M", width=2)
        assign_numeric(mid_group, meta_prefix="M", width=2)
        self.assertTrue(wav_group.records[0].target_name.endswith("_01.wav"))
        self.assertTrue(mid_group.records[0].target_name.endswith("_01.mid"))

    def test_scheme_one_uses_last_three_directories_with_shallow_fallback(self):
        nested = self.root / "A" / "B" / "C"
        nested.mkdir(parents=True)
        (nested / "Deep.wav").write_bytes(b"RIFF")
        self.assertEqual(directory_prefix_defaults(self.root, nested), ("A", "B", "C"))
        self.assertEqual(directory_prefix_defaults(self.root, self.root / "A" / "B"), (self.root.name, "A", "B"))
        self.assertEqual(directory_prefix_defaults(self.root, self.root / "A"), (self.root.name, "A", ""))
        result = scan_folder(self.root)
        deep_group = next(group for group in result.groups.values() if group.relative_folder == "A\\B\\C")
        self.assertEqual((deep_group.meta_prefix, deep_group.prefix, deep_group.records[0].child_prefix), ("A", "B", "C"))

    def test_custom_directory_mapping_and_cross_format_association(self):
        nested = self.root / "Pack" / "Loops" / "Kicks"
        nested.mkdir(parents=True)
        (nested / "Hit_01.wav").write_bytes(b"WAV")
        (nested / "Hit_01.mid").write_bytes(b"MID")
        result = scan_folder(self.root, directory_mapping={"meta": 1, "group": -2, "child": -1})
        wav_group = next(group for group in result.groups.values() if group.relative_folder == "Pack\\Loops\\Kicks" and group.extension == ".wav")
        self.assertEqual((wav_group.meta_prefix, wav_group.prefix, wav_group.records[0].child_prefix), ("Pack", "Loops", "Kicks"))
        self.assertEqual(len(result.associations), 1)
        self.assertEqual(set(result.associations[0]["extensions"]), {".wav", ".mid"})
        self.assertEqual(wav_group.records[0].associated_extensions, [".mid", ".wav"])

    def test_filename_parse_template_and_auto_preview(self):
        if SAMPLE_PACK_WORKFLOW is None:
            self.skipTest("未安装工作流：sample-pack")
        parsed = parse_filename("Loop_Drum_03", "{type}_{name}_{number}")
        self.assertTrue(parsed["matched"])
        self.assertEqual(parsed["fields"]["number"], "03")
        sample_parsed = parse_workflow_filename(
            SAMPLE_PACK_WORKFLOW, "Loop_Drum_03_150BPM", "{type}_{name}_{number}_{bpm}"
        )
        self.assertTrue(sample_parsed["matched"])
        self.assertEqual(sample_parsed["fields"]["bpm"], "150")
        auto = parse_workflow_filename(SAMPLE_PACK_WORKFLOW, "Kick_Loop_150BPM_Am_03")
        self.assertEqual(auto["fields"].get("bpm"), "150")
        self.assertEqual(auto["fields"].get("variant"), "03")

    def test_image_dimensions_and_normalized_metadata(self):
        image = self.root / "Drums&Loop" / "wide.png"
        image.write_bytes(
            b"\x89PNG\r\n\x1a\n" + (13).to_bytes(4, "big") + b"IHDR"
            + (1920).to_bytes(4, "big") + (1080).to_bytes(4, "big")
            + b"\x08\x02\x00\x00\x00"
        )
        self.assertEqual(read_image_dimensions(image), (1920, 1080))
        metadata = read_workflow_metadata(IMAGE_METADATA_WORKFLOW, image, self.root)
        self.assertEqual(metadata["file"]["extension"], "png")
        self.assertEqual(metadata["image"]["orientation"], "landscape")
        self.assertEqual(metadata["image"]["aspect_ratio"], "16:9")
        self.assertEqual(metadata["image"]["aspect_ratio_token"], "16x9")

        portrait = self.root / "Drums&Loop" / "portrait.png"
        portrait.write_bytes(
            b"\x89PNG\r\n\x1a\n" + (13).to_bytes(4, "big") + b"IHDR"
            + (1080).to_bytes(4, "big") + (1920).to_bytes(4, "big")
            + b"\x08\x02\x00\x00\x00"
        )
        scanned = scan_folder(
            self.root,
            metadata_reader=lambda path, root: read_workflow_metadata(IMAGE_METADATA_WORKFLOW, path, root),
        )
        record = next(item for item in scanned.records if item.path == str(portrait))
        self.assertEqual(record.metadata["image"]["orientation"], "portrait")

    def test_complex_image_ratios_snap_to_common_screen_ratios(self):
        cases = {
            "ultrawide.png": (7672, 3264, "21x9", "959x408"),
            "photo.png": (5882, 3892, "3x2", "2941x1946"),
            "dual-monitor.png": (4308, 2137, "2x1", "4308x2137"),
            "portrait.png": (1299, 1813, "3x4", "1299x1813"),
        }
        for filename, (width, height, expected_token, expected_exact) in cases.items():
            image = self.root / "Drums&Loop" / filename
            image.write_bytes(
                b"\x89PNG\r\n\x1a\n" + (13).to_bytes(4, "big") + b"IHDR"
                + width.to_bytes(4, "big") + height.to_bytes(4, "big")
                + b"\x08\x02\x00\x00\x00"
            )
            metadata = read_workflow_metadata(IMAGE_METADATA_WORKFLOW, image, self.root)["image"]
            self.assertEqual(metadata["aspect_ratio_token"], expected_token, filename)
            self.assertEqual(metadata["aspect_ratio_exact"], expected_exact, filename)

    def test_bpm_detection_and_idempotent_suffix(self):
        bpm, source = detect_bpm(self.root / "not-created.wav", "Synth Arp (90, Dm)")
        self.assertEqual((bpm, source), ("90", "name"))
        bpm, source = detect_bpm(self.root / "not-created.mid", "Loop_150")
        self.assertEqual((bpm, source), ("150", "name"))
        wav_fixture = self.root / "metadata-tempo.wav"
        midi_135_fixture = self.root / "metadata-tempo-135.mid"
        midi_140_fixture = self.root / "metadata-tempo-140.mid"
        self._write_bpm_wav(wav_fixture, 150)
        self._write_bpm_midi(midi_135_fixture, 135)
        self._write_bpm_midi(midi_140_fixture, 140)
        bpm, source = detect_bpm(wav_fixture)
        self.assertEqual((bpm, source), ("150", "metadata"))
        bpm, source = detect_bpm(midi_135_fixture)
        self.assertEqual((bpm, source), ("135", "metadata"))
        bpm, source = detect_bpm(midi_140_fixture)
        self.assertEqual((bpm, source), ("140", "metadata"))
        self.assertEqual(append_bpm_suffix("Loop", "150"), "Loop_150BPM")
        self.assertEqual(append_bpm_suffix("Loop_150BPM", "150"), "Loop_150BPM")

    def test_repository_test_resources_are_empty_placeholders(self):
        fixture_root = Path(__file__).resolve().parents[1] / "test-source"
        nonempty = [
            str(path.relative_to(fixture_root))
            for path in fixture_root.rglob("*")
            if path.is_file() and path.stat().st_size
        ]
        self.assertEqual(nonempty, [])

    def test_export_and_sheet_specific_import(self):
        (self.root / "Drums&Loop" / "Tempo_128BPM.wav").write_bytes(b"RIFF")
        output = export_filename_tables(self.root, [".wav", ".png"])
        detail = next(path for path in output if path.name == "Drums&Loop.ori01.ffnf.xlsx")
        workbook = load_workbook(detail, read_only=True, data_only=True)
        self.assertIn("WAV", workbook.sheetnames)
        self.assertIn("Metadata", workbook.sheetnames)
        self.assertIn("Summary", workbook.sheetnames)
        try:
            rows = list(workbook["WAV"].iter_rows(values_only=True))
            self.assertEqual(rows[0][0], "SourceName")
            self.assertIn("Metadata.file.name", rows[0])
            self.assertNotIn("BPM", rows[0])
            tempo_row = next(row for row in rows[1:] if row[0] == "Tempo_128BPM")
        finally:
            workbook.close()
        # A detailed sheet can be selected by extension instead of relying on
        # workbook sheet order.
        result = scan_folder(self.root)
        group = next(group for group in result.groups.values() if group.extension == ".wav")
        match = import_xlsx(detail, group)
        self.assertEqual(match.sheet_name, "WAV")
        self.assertTrue(match.detail_mode)
        self.assertEqual(match.matched_count, 0)

    def test_excel_template_expands_bpm_and_scale_columns(self):
        source = self.root / "Drums&Loop" / "Loop_150BPM.wav"
        source.write_bytes(b"RIFF")
        result = scan_folder(self.root)
        group = next(group for group in result.groups.values()
                     if group.extension == ".wav" and group.folder_name == "Drums&Loop")
        xlsx = self.root / "metadata-mapping.xlsx"
        self._make_workbook(xlsx, [
            ("SourceName", "NewName", "BPM", "Scale"),
            ("Loop_150BPM.wav", "Pad_{scale}_{bpm}", "150", "F#min"),
        ])
        match = import_xlsx(xlsx, group)
        self.assertEqual(match.matched_count, 1)
        self.assertEqual(match.mapping[next(record.path for record in group.records if record.stem == "Loop_150BPM")], "Pad_F#min_150")
        record = next(record for record in group.records if record.stem == "Loop_150BPM")
        self.assertNotIn("scale", record.workflow_values)

        self._make_workbook(xlsx, [
            ("SourceName", "NewName", "BPM", "Scale"),
            ("Loop_150BPM.wav", "Pad_{scale}_{bpm}BPM", "150", ""),
        ])
        empty_scale_match = import_xlsx(xlsx, group)
        self.assertEqual(empty_scale_match.mapping[record.path], "Pad_150BPM")

    def test_export_directory_statistics(self):
        stats = collect_directory_statistics(self.root)
        self.assertGreaterEqual(stats["directory_count"], 3)
        self.assertEqual(stats["file_count"], len(scan_folder(self.root).records))

    def test_undo_rejects_changed_fingerprint(self):
        result = scan_folder(self.root)
        group = next(group for group in result.groups.values() if group.extension == ".mid")
        record = group.records[0]
        record.target_name = "Fingerprint.mid"
        history = self.root / "history.json"
        execute_rename([record], history)
        target = self.root / "MIDI" / "Fingerprint.mid"
        target.write_bytes(b"different")
        ok, errors = undo_last(history)
        self.assertFalse(ok)
        self.assertTrue(any("指纹" in error for error in errors))
        self.assertTrue(target.exists())

    def test_undo_and_redo_resolve_manually_renamed_file_by_fingerprint(self):
        result = scan_folder(self.root)
        group = next(group for group in result.groups.values() if group.extension == ".mid")
        record = group.records[0]
        record.target_name = "FingerprintTarget.mid"
        history = self.root / "history.json"
        execute_rename([record], history)
        target = self.root / "MIDI" / "FingerprintTarget.mid"
        manual = self.root / "MIDI" / "ManuallyChanged.mid"
        target.rename(manual)
        # A same-content replacement at the expected filename must not win
        # over the original file's stable file ID.
        target.write_bytes(manual.read_bytes())

        ok, errors = undo_last(history)
        self.assertTrue(ok, errors)
        original = self.root / "MIDI" / "Pattern1.mid"
        self.assertTrue(original.exists())
        self.assertFalse(manual.exists())
        self.assertTrue(target.exists())

        target.unlink()
        manual_again = self.root / "MIDI" / "ChangedBeforeRedo.mid"
        original.rename(manual_again)
        ok, errors = redo_last(history)
        self.assertTrue(ok, errors)
        self.assertTrue(target.exists())
        self.assertFalse(manual_again.exists())

    def test_undo_rejects_ambiguous_content_fingerprint_candidates(self):
        result = scan_folder(self.root)
        group = next(group for group in result.groups.values() if group.extension == ".mid")
        record = group.records[0]
        record.target_name = "AmbiguousTarget.mid"
        history = self.root / "history.json"
        execute_rename([record], history)

        target = self.root / "MIDI" / "AmbiguousTarget.mid"
        first_candidate = self.root / "MIDI" / "CandidateA.mid"
        second_candidate = self.root / "MIDI" / "CandidateB.mid"
        target.rename(first_candidate)
        second_candidate.write_bytes(first_candidate.read_bytes())

        # Simulate a history entry created on a filesystem where a stable file
        # ID was unavailable, forcing the resolver to use the content identity.
        history_data = json.loads(history.read_text(encoding="utf-8"))
        history_data[-1]["items"][0]["new_fingerprint"]["file_id"] = 0
        history.write_text(json.dumps(history_data, ensure_ascii=False, indent=2), encoding="utf-8")

        ok, errors = undo_last(history)
        self.assertFalse(ok)
        self.assertTrue(any("多个候选" in error for error in errors), errors)
        self.assertTrue(first_candidate.exists())
        self.assertTrue(second_candidate.exists())
        self.assertFalse((self.root / "MIDI" / "Pattern1.mid").exists())

    def test_swap_and_case_only_renames_are_transactional(self):
        second = self.root / "MIDI" / "Pattern2.mid"
        second.write_bytes(b"different-midi")
        result = scan_folder(self.root)
        group = next(group for group in result.groups.values() if group.extension == ".mid")
        first_record = next(record for record in group.records if record.original_name == "Pattern1.mid")
        second_record = next(record for record in group.records if record.original_name == "Pattern2.mid")
        first_record.target_name = "Pattern2.mid"
        second_record.target_name = "Pattern1.mid"
        history = self.root / "history.json"

        operation = execute_rename([first_record, second_record], history)
        self.assertTrue(all(item.success for item in operation.items))
        self.assertEqual((self.root / "MIDI" / "Pattern1.mid").read_bytes(), b"different-midi")
        self.assertEqual((self.root / "MIDI" / "Pattern2.mid").read_bytes(), b"MThd")
        self.assertTrue(undo_last(history)[0])
        self.assertEqual((self.root / "MIDI" / "Pattern1.mid").read_bytes(), b"MThd")
        self.assertEqual((self.root / "MIDI" / "Pattern2.mid").read_bytes(), b"different-midi")

        rescanned = scan_folder(self.root)
        case_group = next(group for group in rescanned.groups.values() if group.extension == ".mid")
        case_record = next(record for record in case_group.records if record.original_name == "Pattern1.mid")
        for record in case_group.records:
            record.selected = record is case_record
        case_record.target_name = "PATTERN1.MID"
        case_history = self.root / "case-history.json"
        case_operation = execute_rename([case_record], case_history)
        self.assertTrue(case_operation.items[0].success)
        self.assertEqual(case_operation.items[0].old_path.lower(), case_operation.items[0].new_path.lower())
        self.assertNotEqual(case_operation.items[0].old_path, case_operation.items[0].new_path)
        self.assertTrue((self.root / "MIDI" / "PATTERN1.MID").exists())
        self.assertTrue(undo_last(case_history)[0])
        self.assertTrue((self.root / "MIDI" / "Pattern1.mid").exists())

    def test_transaction_rolls_back_when_commit_fails(self):
        import namer_core
        root = self.root / "MIDI"
        second = root / "Pattern2.mid"
        second.write_bytes(b"MThd2")
        result = scan_folder(self.root)
        group = next(group for group in result.groups.values() if group.extension == ".mid")
        for index, record in enumerate(group.records, start=1):
            record.target_name = f"Txn_{index}.mid"
        original_rename = namer_core.os.rename
        calls = {"count": 0}
        def fail_commit(source, target):
            calls["count"] += 1
            if calls["count"] == 4:  # stage x2, then fail on second commit
                raise OSError("simulated commit failure")
            return original_rename(source, target)
        operation = None
        with patch.object(namer_core.os, "rename", side_effect=fail_commit):
            operation = execute_rename(group.records, self.root / "history.json")
        self.assertEqual(operation.transaction_status, "rolled_back")
        self.assertTrue((root / "Pattern1.mid").exists())
        self.assertTrue((root / "Pattern2.mid").exists())

    def test_conflict_blocks_group_without_overwrite(self):
        result = scan_folder(self.root)
        group = next(group for group in result.groups.values() if group.extension == ".wav" and group.folder_name == "Drums&Loop")
        group.prefix = ""
        target_record = next(record for record in group.records if record.stem == "Kick 2")
        for record in group.records:
            record.selected = record is target_record
        target_record.name = "existing"
        target_record.target_name = "existing.wav"
        issues = preflight(group.records)
        self.assertTrue(any(issue.code == "exists" for issue in issues))
        history = self.root / "history.json"
        operation = execute_rename(group.records, history)
        self.assertFalse(operation.items[0].success)
        self.assertTrue((self.root / "Drums&Loop" / "existing.wav").exists())

    def test_batch_rename_and_undo(self):
        result = scan_folder(self.root)
        group = next(group for group in result.groups.values() if group.extension == ".mid")
        group.prefix = "MIDI"
        group.records[0].target_name = "MIDI_01.mid"
        history = self.root / "history.json"
        operation = execute_rename(group.records, history)
        self.assertTrue(operation.items[0].success)
        self.assertTrue((self.root / "MIDI" / "MIDI_01.mid").exists())
        ok, errors = undo_last(history)
        self.assertTrue(ok, errors)
        self.assertTrue((self.root / "MIDI" / "Pattern1.mid").exists())

    def test_batch_undo_restores_all_items_and_skips_undone_operation(self):
        second_source = self.root / "MIDI" / "Pattern2.mid"
        second_source.write_bytes(b"MThd")
        result = scan_folder(self.root)
        group = next(group for group in result.groups.values() if group.extension == ".mid")
        for index, record in enumerate(group.records, start=1):
            record.target_name = f"MIDI_{index:02d}.mid"
        history = self.root / "history.json"
        operation = execute_rename(group.records, history)
        self.assertEqual(len(operation.items), 2)
        self.assertTrue(all(item.success for item in operation.items))
        ok, errors = undo_last(history)
        self.assertTrue(ok, errors)
        self.assertTrue((self.root / "MIDI" / "Pattern1.mid").exists())
        self.assertTrue((self.root / "MIDI" / "Pattern2.mid").exists())
        ok_again, errors_again = undo_last(history)
        self.assertFalse(ok_again)
        self.assertIn("没有可撤销", errors_again[0])

    def test_redo_restores_all_items_and_can_be_undone_again(self):
        second_source = self.root / "MIDI" / "Pattern2.mid"
        second_source.write_bytes(b"MThd")
        result = scan_folder(self.root)
        group = next(group for group in result.groups.values() if group.extension == ".mid")
        for index, record in enumerate(group.records, start=1):
            record.target_name = f"Redo_{index:02d}.mid"
        history = self.root / "history.json"
        execute_rename(group.records, history)
        self.assertTrue(undo_last(history)[0])
        ok, errors = redo_last(history)
        self.assertTrue(ok, errors)
        self.assertTrue((self.root / "MIDI" / "Redo_01.mid").exists())
        self.assertTrue((self.root / "MIDI" / "Redo_02.mid").exists())
        data = json.loads(history.read_text(encoding="utf-8"))
        self.assertNotIn("undone_at", data[0])
        self.assertTrue(all(not item.get("undone") for item in data[0]["items"]))
        ok_again, errors_again = undo_last(history)
        self.assertTrue(ok_again, errors_again)
        self.assertTrue(second_source.exists())
        self.assertTrue((self.root / "MIDI" / "Pattern1.mid").exists())

    def test_redo_does_not_overwrite_existing_target(self):
        result = scan_folder(self.root)
        group = next(group for group in result.groups.values() if group.extension == ".mid")
        record = group.records[0]
        record.target_name = "RedoConflict.mid"
        history = self.root / "history.json"
        execute_rename(group.records, history)
        self.assertTrue(undo_last(history)[0])
        target = self.root / "MIDI" / "RedoConflict.mid"
        target.write_bytes(b"user file")
        ok, errors = redo_last(history)
        self.assertFalse(ok)
        self.assertTrue(any("已存在" in error for error in errors))
        self.assertTrue(record.source_path.exists())
        self.assertEqual(target.read_bytes(), b"user file")

    def test_redo_uses_most_recent_undo_order(self):
        first = self.root / "MIDI" / "Pattern1.mid"
        second = self.root / "MIDI" / "Pattern2.mid"
        second.write_bytes(b"MThd")
        result = scan_folder(self.root)
        group = next(group for group in result.groups.values() if group.extension == ".mid")
        history = self.root / "history.json"
        group.records[0].target_name = "FirstRedo.mid"
        execute_rename([group.records[0]], history, kind="single")
        group.records[1].target_name = "SecondRedo.mid"
        execute_rename([group.records[1]], history, kind="single")
        self.assertTrue(undo_last(history)[0])
        self.assertTrue(undo_last(history)[0])
        self.assertTrue(redo_last(history)[0])
        self.assertTrue((self.root / "MIDI" / "FirstRedo.mid").exists())
        self.assertTrue(second.exists())
        self.assertFalse(first.exists())

    def test_undo_groups_legacy_entries_from_one_all_groups_action(self):
        result = scan_folder(self.root)
        wav_group = next(group for group in result.groups.values() if group.extension == ".wav" and group.folder_name == "Drums&Loop")
        mid_group = next(group for group in result.groups.values() if group.extension == ".mid")
        wav_record = next(record for record in wav_group.records if record.stem == "Kick 2")
        mid_record = mid_group.records[0]
        wav_record.selected = True
        mid_record.selected = True
        wav_record.target_name = "LegacyKick.wav"
        mid_record.target_name = "LegacyPattern.mid"
        for record in wav_group.records:
            if record is not wav_record:
                record.selected = False
        history = self.root / "history.json"
        execute_rename(wav_group.records, history)
        execute_rename(mid_group.records, history)
        data = json.loads(history.read_text(encoding="utf-8"))
        data[-1]["operation_time"] = data[-2]["operation_time"]
        history.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
        ok, errors = undo_last(history)
        self.assertTrue(ok, errors)
        self.assertTrue((self.root / "Drums&Loop" / "Kick 2.wav").exists())
        self.assertTrue((self.root / "MIDI" / "Pattern1.mid").exists())

    def _make_workbook(self, path: Path, rows, two_columns=True):
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row if two_columns else [row[0]])
        wb.save(path)

    def test_excel_source_and_name_mode(self):
        result = scan_folder(self.root)
        group = next(group for group in result.groups.values() if group.extension == ".wav" and group.folder_name == "Drums&Loop")
        xlsx = self.root / "mapping.xlsx"
        self._make_workbook(xlsx, [("Kick 2.WAV", "Chest"), ("not-there.wav", "Nope"), ("existing.wav", "")])
        match = import_xlsx(xlsx, group)
        self.assertEqual(match.mode, "source-and-name")
        self.assertEqual(match.matched_count, 1)
        kick = next(record for record in group.records if record.stem == "Kick 2")
        self.assertEqual(match.mapping[kick.path], "Chest")
        self.assertEqual(len(match.unmatched_rows), 1)
        self.assertEqual(len(match.unmatched_files), 1)
        self.assertEqual(len(match.matched_without_name), 1)

    def test_excel_ordered_mode_and_extra_rows(self):
        result = scan_folder(self.root)
        group = next(group for group in result.groups.values() if group.extension == ".mid")
        xlsx = self.root / "ordered.xlsx"
        self._make_workbook(xlsx, [("First",), ("",), ("Extra",)], two_columns=False)
        match = import_xlsx(xlsx, group)
        self.assertEqual(match.mode, "ordered-names")
        self.assertEqual(match.matched_count, 1)
        self.assertTrue(any("为空" in warning for warning in match.warnings))
        self.assertTrue(any("超出" in warning for warning in match.warnings))

    def test_export_tables_collision_and_generated_skip(self):
        # Existing output exercises .ori01; export only selected formats.
        (self.root / "Drums&Loop.ffnf.xlsx").write_bytes(b"existing")
        outputs = export_filename_tables(self.root, [".wav", ".png"], include_hidden=False)
        names = {path.name for path in outputs}
        self.assertIn("Drums&Loop.ori01.ffnf.xlsx", names)
        output = next(path for path in outputs if path.name.startswith("Drums&Loop"))
        wb = load_workbook(output, read_only=True)
        self.assertIn("WAV", wb.sheetnames)
        self.assertIn("PNG", wb.sheetnames)
        values = [row[0] for row in wb["WAV"].iter_rows(values_only=True)][1:]
        self.assertEqual(values, sorted(values, key=natural_key))
        wb.close()
        # A second export must advance to ori02 and ignore the first generated sheet.
        outputs2 = export_filename_tables(self.root, [".wav", ".png"])
        self.assertTrue(any(path.name == "Drums&Loop.ori02.ffnf.xlsx" for path in outputs2))

    def test_export_deep_tree_writes_filetree_once_without_overwrite(self):
        nested = self.root / "Level 1" / "Level 2" / "Level 3"
        nested.mkdir(parents=True)
        (nested / "Deep 2.wav").write_bytes(b"RIFF")
        (self.root / "Level 1" / "Level 2" / "Empty Level 3").mkdir()
        outputs = export_filename_tables(self.root, [".wav"])
        filetree = self.root / "filetree.txt"
        self.assertIn(filetree, outputs)
        filetree_text = filetree.read_text(encoding="utf-8").replace("\\", "/")
        self.assertIn("Level 1/Level 2/Level 3", filetree_text)
        self.assertIn("Level 1/Level 2/Level 3.ffnf.xlsx", filetree_text)
        self.assertNotIn("Empty Level 3", filetree_text)
        self.assertNotIn("Deep 2.wav", filetree_text)
        original = filetree_text
        outputs_again = export_filename_tables(self.root, [".wav"])
        self.assertNotIn(filetree, outputs_again)
        self.assertEqual(filetree.read_text(encoding="utf-8").replace("\\", "/"), original)
        self.assertNotIn(".txt", scan_folder(self.root).extension_counts)


if __name__ == "__main__":
    unittest.main()
